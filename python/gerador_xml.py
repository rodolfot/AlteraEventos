"""
Gerador de XML a partir de Planilhas Excel/CSV
Interface Tkinter completa - Python 3.8+

Funcionalidades:
  - Carregar planilha principal (aba "Campos Entrada")
  - Carregar planilha origem e copiar campos para a principal
  - Tabela editável com Campo/PosIni/PosFin/Tamanho/Tipo/Alinhamento/Valor
  - Validação da soma de tamanhos x posições
  - Geração de XML com atributos posicionados corretamente
  - Salvar planilha atualizada de volta em .xlsx ou .csv
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import xml.etree.ElementTree as ET
from xml.dom import minidom
import openpyxl
import csv
import os
import re
import threading
import time
import shutil
import html
from openpyxl.utils import get_column_letter
import unicodedata


# ─────────────────────────────────────────────────────────────────────────────
# Constantes de cor / estilo
# ─────────────────────────────────────────────────────────────────────────────
COR_BG          = "#f4f6f8"
COR_TOOLBAR     = "#eceff1"
COR_BTN_VERDE   = "#43a047"
COR_BTN_AZUL    = "#1e88e5"
COR_BTN_LARANJA = "#fb8c00"
COR_BTN_ROXO    = "#8e24aa"
COR_BTN_CINZA   = "#546e7a"
COR_BTN_TEAL    = "#00897b"
COR_BTN_VERM    = "#e53935"
COR_BRANCO      = "#ffffff"
FONT_NORMAL     = ("Segoe UI", 9)
FONT_BOLD       = ("Segoe UI", 9, "bold")
FONT_MONO       = ("Consolas", 9)


# ─────────────────────────────────────────────────────────────────────────────
# Utilitários de leitura de planilha
# ─────────────────────────────────────────────────────────────────────────────

def _normalizar_chave(texto):
    """Normaliza texto para comparação: minúsculo, sem espaços/underscores."""
    return re.sub(r"[\s_\-]", "", str(texto or "")).lower()


def _cell_str(valor, default=""):
    """Converte valor de célula para string limpa."""
    if valor is None:
        return default
    if isinstance(valor, float):
        return str(int(valor)) if valor == int(valor) else str(valor)
    return str(valor).strip()


def _cell_int(valor):
    """Converte valor de célula para int ou None."""
    try:
        return int(float(str(valor).strip()))
    except (ValueError, TypeError):
        return None


def _detectar_linha_cabecalho(sheet):
    """Localiza a linha do cabeçalho procurando 'NomeCampo' nas 10 primeiras linhas."""
    alvo = {"nomecampo", "nome", "campo", "fieldname"}
    for row_idx in range(1, min(11, sheet.max_row + 1)):
        for cell in sheet[row_idx]:
            if _normalizar_chave(cell.value) in alvo:
                return row_idx
    return 2  # padrão: linha 2


def _mapear_colunas(sheet, header_row):
    """Retorna {chave_normalizada: índice_coluna_1based} da linha de cabeçalho."""
    mapa = {}
    for cell in sheet[header_row]:
        if cell.value:
            mapa[_normalizar_chave(cell.value)] = cell.column
    return mapa


def _get_col(row_cells, col_map, *chaves, default=""):
    """Lê célula pelo nome normalizado da coluna."""
    for chave in chaves:
        col = col_map.get(_normalizar_chave(chave))
        if col:
            val = _cell_str(row_cells[col - 1].value, default)
            if val != default or default != "":
                return val
    return default


def ler_campos_entrada(filepath):
    """
    Lê a aba 'Campos Entrada' de um .xlsx ou qualquer aba de um .csv.
    Retorna lista de dicts com os campos do evento.
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext in (".xlsx", ".xls"):
        return _ler_xlsx_campos_entrada(filepath)
    elif ext == ".csv":
        return _ler_csv_campos_entrada(filepath)
    else:
        raise ValueError(f"Formato não suportado: {ext}")


def _detectar_secoes(sheet, meta_row_idx):
    """
    Lê a linha de metadados (acima do cabeçalho) e retorna {nome_secao: [col_names]}.
    Usado para delimitar quais colunas pertencem a cada seção XML
    (ex.: 'Layouts', 'Campos', 'Layout Entrada', 'LayoutPersistencia', ...).
    """
    meta_cells = []
    for cell in sheet[meta_row_idx]:
        if cell.value and str(cell.value).strip():
            meta_cells.append((cell.column, str(cell.value).strip()))
    if not meta_cells:
        return {}

    # Mapeamento coluna → nome de header na linha seguinte (cabeçalho real)
    header_row_idx = meta_row_idx + 1
    col_to_header = {
        cell.column: str(cell.value).strip()
        for cell in sheet[header_row_idx]
        if cell.value
    }

    sections = {}
    for i, (sec_col, sec_name) in enumerate(meta_cells):
        next_col = meta_cells[i + 1][0] if i + 1 < len(meta_cells) else float("inf")
        sec_headers = [
            col_to_header[c]
            for c in sorted(col_to_header)
            if sec_col <= c < next_col
        ]
        sections[sec_name] = sec_headers
    return sections


def _ler_campos_de_sheet(sheet):
    """
    Lê campos de qualquer aba de planilha, detectando cabeçalho e colunas automaticamente.
    Retorna (campos, headers, sections) onde:
      - campos: lista de dicts com chaves padrão + '_raw' (todos os valores pelo header original)
      - headers: lista de nomes de colunas na ordem original da planilha
      - sections: {nome_secao: [col_names]} detectado da linha de metadados (se existir)
    """
    header_row = _detectar_linha_cabecalho(sheet)
    col_map = _mapear_colunas(sheet, header_row)

    # Cabeçalhos na ordem original da planilha
    headers = []
    col_para_header = {}
    for cell in sheet[header_row]:
        if cell.value:
            h = str(cell.value).strip()
            headers.append(h)
            col_para_header[cell.column] = h

    campos = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        row = sheet[row_idx]

        nome = _get_col(row, col_map, "NomeCampo", "Nome", "Campo")
        if not nome:
            continue

        tamanho = _cell_int(_get_col(row, col_map, "TamanhoCampo", "Tamanho"))
        pos_ini = _cell_int(_get_col(row, col_map, "PosicaoInicial", "PosInicial", "PosIni"))
        pos_fin_lido = _cell_int(_get_col(row, col_map, "PosicaoFinal", "PosFinal", "PosFin"))
        pos_fin = (pos_ini + tamanho - 1) if (pos_ini and tamanho) else pos_fin_lido

        # Todos os valores da linha pelo nome original do cabeçalho
        raw = {}
        for cell in row:
            if cell.column in col_para_header:
                raw[col_para_header[cell.column]] = _cell_str(cell.value)

        campo = {
            "linha":      row_idx,
            "entrada":    _get_col(row, col_map, "Entrada", default="S"),
            "id":         _get_col(row, col_map, "IdentificadorCampo", "ID", "Id"),
            "nome":       nome,
            "descricao":  _get_col(row, col_map, "DescricaoCampo", "Descricao"),
            "tipo":       _get_col(row, col_map, "TipoCampo", "Tipo", default="TEXTO"),
            "tamanho":    tamanho,
            "pos_ini":    pos_ini,
            "pos_fin":    pos_fin,
            "valor_padrao": _get_col(row, col_map, "ValorPadrao", "Valor_Padrao"),
            "alinhamento":  _get_col(row, col_map, "AlinhamentoCampo", "Alinhamento"),
            "obrigatorio":  _get_col(row, col_map, "CampoObrigatorio", "Obrigatorio"),
            "coluna_db":    _get_col(row, col_map, "NomeColuna", "Coluna_DB", "ColunaDB"),
            "oracle_type":  _get_col(row, col_map, "OracleDataType", "OracleType"),
            "valor":        _get_col(row, col_map, "ValorPadrao", "Valor_Padrao"),
            "_raw":         raw,
        }
        campos.append(campo)

    # Detecta seções da linha de metadados acima do cabeçalho (se existir)
    sections = _detectar_secoes(sheet, header_row - 1) if header_row > 1 else {}

    return campos, headers, sections


def _ler_xlsx_campos_entrada(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Localiza a aba
    sheet = None
    for name in wb.sheetnames:
        if _normalizar_chave(name) in ("camposentrada", "campos_entrada", "campos entrada"):
            sheet = wb[name]
            break
    if sheet is None:
        raise ValueError(
            f"Aba 'Campos Entrada' não encontrada.\n"
            f"Abas disponíveis: {', '.join(wb.sheetnames)}"
        )

    campos, _, _ = _ler_campos_de_sheet(sheet)
    return campos


def _ler_csv_campos_entrada(filepath):
    campos = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            norm = {_normalizar_chave(k): v.strip() for k, v in row.items()}
            nome = norm.get("nomecampo") or norm.get("nome") or ""
            if not nome:
                continue

            tamanho = _cell_int(norm.get("tamanhocampo") or norm.get("tamanho"))
            pos_ini = _cell_int(norm.get("posicaoinicial") or norm.get("posinicial"))
            pos_fin = (pos_ini + tamanho - 1) if (pos_ini and tamanho) else None

            campo = {
                "linha":      i + 2,
                "entrada":    norm.get("entrada", "S"),
                "id":         norm.get("identificadorcampo") or norm.get("id", ""),
                "nome":       nome,
                "descricao":  norm.get("descricaocampo") or norm.get("descricao", ""),
                "tipo":       norm.get("tipocampo") or norm.get("tipo", "TEXTO"),
                "tamanho":    tamanho,
                "pos_ini":    pos_ini,
                "pos_fin":    pos_fin,
                "valor_padrao": norm.get("valorpadrao", ""),
                "alinhamento":  norm.get("alinhamentocampo") or norm.get("alinhamento", ""),
                "obrigatorio":  norm.get("campoobrigatorio") or norm.get("obrigatorio", ""),
                "coluna_db":    norm.get("nomecoluna") or norm.get("colunadb", ""),
                "oracle_type":  norm.get("oracledatatype", ""),
                "valor":        norm.get("valorpadrao", ""),
            }
            campos.append(campo)
    return campos


def ler_todas_abas(filepath):
    """
    Lê todas as abas de um .xlsx como dicionário {nome_aba: {"campos": list, "headers": list}}.
    Para CSV, retorna uma única entrada 'Campos Entrada'.
    Abas sem campos reconhecidos são ignoradas.
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".csv":
        campos = _ler_csv_campos_entrada(filepath)
        return {"Campos Entrada": {"campos": campos, "headers": []}}

    wb = openpyxl.load_workbook(filepath, data_only=True)
    resultado = {}

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        try:
            campos, headers, sections = _ler_campos_de_sheet(ws)
            if campos:
                resultado[nome_aba] = {"campos": campos, "headers": headers, "sections": sections}
        except Exception:
            pass

    return resultado


def _nome_xml_para_aba(nome_aba):
    """Converte nome de aba em nome do elemento XML raiz. Ex: 'Campos Entrada' → 'LayoutEntrada'."""
    secao = re.sub(r"^[Cc]ampos\s+", "", nome_aba).strip()
    pascal = "".join(p.capitalize() for p in re.split(r"[\s_\-]+", secao) if p)
    return f"Layout{pascal}" if pascal else "Layout"


def _item_xml_para_aba(nome_aba):
    """Converte nome de aba em nome do elemento XML de campo. Ex: 'Campos Entrada' → 'CampoEntrada'."""
    secao = re.sub(r"^[Cc]ampos\s+", "", nome_aba).strip()
    pascal = "".join(p.capitalize() for p in re.split(r"[\s_\-]+", secao) if p)
    return f"Campo{pascal}" if pascal else "Campo"


def salvar_xlsx(filepath, campos, nome_aba="Campos Entrada"):
    """Salva lista de campos na aba indicada do arquivo Excel."""
    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception:
        wb = openpyxl.Workbook()
        wb.active.title = nome_aba

    sheet_name = nome_aba
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]

    # Cabeçalho na linha 2 (padrão do template)
    headers = [
        "Entrada", "Persistencia", "Enriquecimento", "MapaAtributo", "Saida", "CampoConcatenado",
        "IdentificadorCampo", "NomeCampo", "DescricaoCampo", "TipoCampo", "TamanhoCampo",
        "PosicaoInicial", "PosicaoFinal", "ValorPadrao", "AlinhamentoCampo", "CampoObrigatorio",
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=2, column=col_idx, value=header)

    # Dados a partir da linha 6
    for i, campo in enumerate(campos):
        r = 6 + i
        ws.cell(r, 1,  campo.get("entrada", "S"))
        ws.cell(r, 7,  campo.get("id") or i + 1)
        ws.cell(r, 8,  campo.get("nome", ""))
        ws.cell(r, 9,  campo.get("descricao", ""))
        ws.cell(r, 10, campo.get("tipo", "TEXTO"))
        ws.cell(r, 11, campo.get("tamanho"))
        ws.cell(r, 12, campo.get("pos_ini"))
        if campo.get("pos_ini") and campo.get("tamanho"):
            ws.cell(r, 13).value = f"=L{r}+K{r}-1"
        ws.cell(r, 14, campo.get("valor_padrao", ""))
        ws.cell(r, 15, campo.get("alinhamento", ""))
        ws.cell(r, 16, campo.get("obrigatorio", ""))

    wb.save(filepath)


def salvar_csv(filepath, campos):
    """Salva lista de campos em CSV."""
    fieldnames = ["entrada", "id", "nome", "descricao", "tipo", "tamanho",
                  "pos_ini", "pos_fin", "valor_padrao", "alinhamento",
                  "obrigatorio", "coluna_db", "valor"]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(campos)


def _to_cell_value(v):
    """Converte string de _raw para o tipo adequado para escrita em célula xlsx."""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(s)
    except (ValueError, TypeError):
        pass
    try:
        return float(s)
    except (ValueError, TypeError):
        pass
    return s


def salvar_xlsx_estruturado(path_original, path_destino, dados_por_aba):
    """
    Persiste as alterações de dados_por_aba no xlsx preservando toda a estrutura original.

    Estratégia:
      1. Copia o arquivo original para path_destino (todos os tabs, formatação, imagens, etc.)
      2. Para cada aba presente em dados_por_aba:
           - Detecta a linha de cabeçalho e a primeira linha de dados no original
           - Limpa apenas as células de dados (preserva linhas de título/seção acima)
           - Reescreve todos os campos usando _raw (header original → número da coluna)
           - PosicaoFinal é escrita como fórmula Excel (=PosIni+Tam-1)
      3. Abas não presentes em dados_por_aba são intocadas.
      4. O arquivo original NUNCA é modificado.
    """
    shutil.copy2(path_original, path_destino)
    wb = openpyxl.load_workbook(path_destino)

    for nome_aba, info in dados_por_aba.items():
        if nome_aba not in wb.sheetnames:
            continue

        campos = info.get("campos", [])
        ws = wb[nome_aba]

        # ── Detectar cabeçalho ──────────────────────────────────────────────
        header_row = _detectar_linha_cabecalho(ws)

        # Mapa: header_name_original → col_number
        col_map = {}
        for cell in ws[header_row]:
            if cell.value:
                col_map[str(cell.value).strip()] = cell.column

        if not col_map:
            continue

        # ── Detectar primeira linha de dados reais ──────────────────────────
        # Percorre até 20 linhas após o cabeçalho procurando a primeira com dado
        primeira_linha = header_row + 1
        for r in range(header_row + 1, min(header_row + 21, ws.max_row + 1)):
            if any(ws.cell(r, c).value is not None for c in col_map.values()):
                primeira_linha = r
                break

        # ── Colunas especiais para fórmula PosicaoFinal ─────────────────────
        col_posini = col_map.get("PosicaoInicial") or col_map.get("PosInicial")
        col_tam    = col_map.get("TamanhoCampo")   or col_map.get("Tamanho")
        col_posfin = col_map.get("PosicaoFinal")   or col_map.get("PosFinal")

        # ── Limpar linhas de dados (só as colunas mapeadas) ─────────────────
        for r in range(primeira_linha, ws.max_row + 1):
            for col_num in col_map.values():
                ws.cell(r, col_num).value = None

        # ── Reescrever campos usando _raw ────────────────────────────────────
        for i, campo in enumerate(campos):
            r = primeira_linha + i
            raw = campo.get("_raw", {})

            for header_name, col_num in col_map.items():
                # PosicaoFinal: fórmula Excel para manter a referência dinâmica
                if (header_name in ("PosicaoFinal", "PosFinal")
                        and col_posini and col_tam and col_posfin):
                    letra_ini = get_column_letter(col_posini)
                    letra_tam = get_column_letter(col_tam)
                    ws.cell(r, col_num).value = f"={letra_ini}{r}+{letra_tam}{r}-1"
                    continue

                ws.cell(r, col_num).value = _to_cell_value(raw.get(header_name))

    wb.save(path_destino)


# ─────────────────────────────────────────────────────────────────────────────
# XML
# ─────────────────────────────────────────────────────────────────────────────

def _indent_et(elem, level=0):
    """
    Indenta um ElementTree in-place para pretty-print (Python 3.8 compatible).
    Substitui minidom.toprettyxml para XML com prefixos de namespace,
    evitando o erro "prefix format reserved for internal use" do expat.
    """
    pad = "\n" + "\t" * level
    if len(elem):
        if not (elem.text or "").strip():
            elem.text = pad + "\t"
        for i, child in enumerate(elem):
            _indent_et(child, level + 1)
            is_last = (i == len(elem) - 1)
            if not (child.tail or "").strip():
                child.tail = pad if is_last else pad + "\t"
    else:
        if not (elem.text or "").strip():
            elem.text = ""


def _sanitizar_xml(nome):
    s = re.sub(r"[^a-zA-Z0-9_\-.]", "_", (nome or "campo").strip())
    if s and not (s[0].isalpha() or s[0] == "_"):
        s = "_" + s
    return s or "campo"


def _aplicar_alinhamento(valor, tamanho, alinhamento, tipo):
    valor = str(valor or "")
    if len(valor) > tamanho:
        return valor[:tamanho]
    diff = tamanho - len(valor)
    if diff == 0:
        return valor

    alin = (alinhamento or "").upper().strip()
    if not alin:
        numerico = any(t in (tipo or "").upper() for t in ("INTEIRO", "DECIMAL", "NUMERO"))
        alin = "ZERO_ESQUERDA" if numerico else "BRANCO_ESQUERDA"

    if alin == "BRANCO_DIREITA":   return " " * diff + valor
    if alin == "BRANCO_ESQUERDA":  return valor + " " * diff
    if alin == "ZERO_ESQUERDA":    return "0" * diff + valor
    if alin == "ZERO_DIREITA":     return valor + "0" * diff
    return valor + " " * diff


def construir_xml(campos, headers=None, nome_aba="", sections=None):
    """
    Constrói string XML no formato Layout* (gabarito) a partir dos campos da aba.

    Estrutura gerada:
      <LayoutEntrada>          ← derivado de nome_aba via _nome_xml_para_aba
        <Campos>
          <CampoEntrada>       ← derivado de nome_aba via _item_xml_para_aba
            <IdentificadorCampo>...
            ...
            <Posicao>
              <PosicaoInicial>...
              <PosicaoFinal>...
            </Posicao>
          </CampoEntrada>
        </Campos>
      </LayoutEntrada>

    Parâmetros:
      headers  - lista de nomes de colunas originais da planilha (ordem natural)
      nome_aba - nome da aba (determina root/item element names)
      sections - {nome_secao: [col_names]} lido da linha de metadados da planilha
    """
    # ── Nomes dos elementos XML ───────────────────────────────────────────────
    root_tag = _nome_xml_para_aba(nome_aba) if nome_aba else "Layout"
    item_tag = _item_xml_para_aba(nome_aba) if nome_aba else "Campo"
    cont_tag = "Campos"

    # ── Campos ativos (Entrada = S, com posição definida) ─────────────────────
    ativos = sorted(
        [c for c in campos if (c.get("entrada", "S") or "S").upper() == "S" and c.get("pos_ini")],
        key=lambda c: c.get("pos_ini", 0)
    )

    # ── Colunas de posição (vão aninhadas em <Posicao>) ───────────────────────
    _POS = frozenset(["posicaoinicial", "posinicial", "posicaofinal", "posfinal"])

    # ── Determinar quais headers emitir e quais são de posição ────────────────
    if headers and sections:
        # Colunas de flag de layout (ex.: Entrada, Persistência, Saída…)
        flag_cols = set(sections.get("Layouts", []))

        # Colunas compartilhadas (seção "Campos")
        shared = sections.get("Campos", [])

        # Colunas específicas da seção correspondente a esta aba
        #  "Campos Entrada" → busca seção "Layout Entrada" ou "LayoutEntrada"
        secao_pascal = re.sub(r"^[Cc]ampos\s+", "", nome_aba).strip()
        secao_pascal = "".join(p.capitalize() for p in re.split(r"[\s_\-]+", secao_pascal) if p)
        aba_sec = next(
            (s for s in sections
             if re.sub(r"\s+", "", s).lower() == f"layout{secao_pascal.lower()}"),
            None
        )
        specific = sections.get(aba_sec, []) if aba_sec else []

        include_all = [h for h in (shared + specific) if h not in flag_cols]

    elif headers:
        # Sem metadados de seção: exclui flags conhecidas pelo nome
        _FLAG_KNOWN = frozenset([
            "entrada", "persistência", "persistencia",
            "enriquecimento", "mapaatributo",
            "saída", "saida", "campoconcatenado",
        ])
        include_all = [h for h in headers if _normalizar_chave(h) not in _FLAG_KNOWN]
    else:
        include_all = []

    headers_pos = [h for h in include_all if _normalizar_chave(h) in _POS]
    headers_xml = [h for h in include_all if _normalizar_chave(h) not in _POS]

    # ── Construção do XML ─────────────────────────────────────────────────────
    root_el = ET.Element(root_tag)
    cont_el = ET.SubElement(root_el, cont_tag)

    for c in ativos:
        raw     = c.get("_raw", {})
        item_el = ET.SubElement(cont_el, item_tag)

        if headers_xml or headers_pos:
            # Emite colunas principais (sem flags, sem posição) em ordem
            for h in headers_xml:
                val = raw.get(h, "")
                if val:
                    e = ET.SubElement(item_el, _sanitizar_xml(h))
                    e.text = str(val)

            # Emite <Posicao> com PosicaoInicial e PosicaoFinal
            pos_vals = [(h, raw.get(h, "")) for h in headers_pos]
            if any(v for _, v in pos_vals):
                pos_el = ET.SubElement(item_el, "Posicao")
                for h, v in pos_vals:
                    if v:
                        e = ET.SubElement(pos_el, _sanitizar_xml(h))
                        e.text = str(v)
        else:
            # Fallback sem headers: usa campos processados
            def _sub(tag, val, _el=item_el):
                if val is not None and str(val).strip():
                    e = ET.SubElement(_el, tag)
                    e.text = str(val)

            _sub("IdentificadorCampo", c.get("id", ""))
            _sub("NomeCampo",          c.get("nome", ""))
            _sub("DescricaoCampo",     c.get("descricao", ""))
            _sub("TipoCampo",          c.get("tipo", ""))
            _sub("TamanhoCampo",       str(c["tamanho"]) if c.get("tamanho") else "")
            _sub("AlinhamentoCampo",   c.get("alinhamento", ""))
            _sub("CampoObrigatorio",   c.get("obrigatorio", ""))

            pos_ini = c.get("pos_ini")
            pos_fin = c.get("pos_fin") or (pos_ini + c["tamanho"] - 1 if pos_ini and c.get("tamanho") else None)
            if pos_ini or pos_fin:
                pos_el = ET.SubElement(item_el, "Posicao")
                if pos_ini:
                    e = ET.SubElement(pos_el, "PosicaoInicial"); e.text = str(pos_ini)
                if pos_fin:
                    e = ET.SubElement(pos_el, "PosicaoFinal");   e.text = str(pos_fin)

    raw_xml = ET.tostring(root_el, encoding="unicode")
    return minidom.parseString(raw_xml).toprettyxml(indent="\t")


# ─────────────────────────────────────────────────────────────────────────────
# Geradores XML específicos (LayoutPersistencia, MapaAtributo, Enriquecimento)
# ─────────────────────────────────────────────────────────────────────────────

def _norm_aba(nome):
    """Normaliza nome de aba removendo acentos, espaços e convertendo para minúsculas."""
    texto = unicodedata.normalize("NFKD", str(nome or ""))
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return re.sub(r"[\s_\-]", "", texto).lower()


def _raw_flag(raw, *keys):
    """Retorna True se alguma das chaves (tolerando acentos) tiver valor 'S' no dict raw."""
    raw_norm = {_norm_aba(k): v for k, v in raw.items()}
    for k in keys:
        if str(raw_norm.get(_norm_aba(k), "")).strip().upper() == "S":
            return True
    return False


def _aba_campos_entrada(dados_por_aba):
    """Retorna a lista de campos da aba 'Campos Entrada'."""
    for nome, info in dados_por_aba.items():
        if _norm_aba(nome) in ("camposentrada", "camposdeentrada"):
            return info.get("campos", [])
    if dados_por_aba:
        return next(iter(dados_por_aba.values())).get("campos", [])
    return []


def _ler_identificacao_evento(filepath):
    """
    Lê a aba 'Identificação Evento' diretamente do xlsx.
    Retorna dict {header: valor} da primeira linha de dados.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = None
        for name in wb.sheetnames:
            n = _norm_aba(name)
            if "identificacaoevento" in n or "identificaevento" in n:
                sheet = wb[name]
                break
        if sheet is None:
            return {}
        header_row = _detectar_linha_cabecalho(sheet)
        headers_map = {}
        for cell in sheet[header_row]:
            if cell.value:
                headers_map[cell.column] = str(cell.value).strip()
        for row_idx in range(header_row + 1, min(header_row + 5, sheet.max_row + 1)):
            row = sheet[row_idx]
            result = {}
            for cell in row:
                if cell.column in headers_map and cell.value is not None:
                    result[headers_map[cell.column]] = _cell_str(cell.value)
            if result:
                return result
    except Exception:
        pass
    return {}


def _ler_rule_attribute_valores(filepath):
    """
    Lê a aba 'Rule Attribute Valor Padrão' do xlsx.
    Retorna lista de dicts com as colunas da aba (dataType, value, pattern etc.).
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = None
        for name in wb.sheetnames:
            n = _norm_aba(name)
            if "ruleattribute" in n and "valor" in n:
                sheet = wb[name]
                break
        if sheet is None:
            return []
        headers_map = {}
        for cell in sheet[1]:
            if cell.value:
                headers_map[cell.column] = str(cell.value).strip()
        result = []
        for row_idx in range(2, sheet.max_row + 1):
            row = sheet[row_idx]
            item = {}
            for cell in row:
                if cell.column in headers_map and cell.value is not None:
                    item[headers_map[cell.column]] = _cell_str(cell.value)
            if item:
                result.append(item)
        return result
    except Exception:
        pass
    return []


def construir_xml_persistencia(dados_por_aba, filepath=None):
    """
    Gera XML LayoutPersistencia a partir dos campos com Persistência=S.
    Metadados de cabeçalho (Identificador, TamanhoLayout, IdentificadorEvento)
    lidos da aba 'Identificação Evento' diretamente do xlsx.

    Estrutura:
      <LayoutPersistencia>
        <Identificador>...</Identificador>
        <TamanhoLayout>...</TamanhoLayout>
        <IdentificadorEvento>...</IdentificadorEvento>
        <Campos>
          <CampoPersistencia>
            <NomeTabela>...</NomeTabela>
            <NomeColuna>...</NomeColuna>
            <ValorPadrao>...</ValorPadrao>   ← apenas se não vazio
            <AlinhamentoCampo>...</AlinhamentoCampo>
            <IdentificadorCampo>...</IdentificadorCampo>
            <NomeCampo>...</NomeCampo>
            <DescricaoCampo>...</DescricaoCampo>
            <TipoCampo>...</TipoCampo>
            <CampoObrigatorio>...</CampoObrigatorio>
            <TamanhoCampo>...</TamanhoCampo>  ← apenas se não vazio
          </CampoPersistencia>
        </Campos>
      </LayoutPersistencia>
    """
    campos_entrada = _aba_campos_entrada(dados_por_aba)
    campos_pers = [
        c for c in campos_entrada
        if _raw_flag(c.get("_raw", {}), "Persistência", "Persistencia")
    ]

    id_evento = _ler_identificacao_evento(filepath) if filepath else {}
    id_norm = {_norm_aba(k): v for k, v in id_evento.items()}

    root_el = ET.Element("LayoutPersistencia")

    identificador = id_norm.get("identificador", "")
    if identificador:
        ET.SubElement(root_el, "Identificador").text = identificador

    # TamanhoLayout = PosicaoFinal do último campo de Campos Entrada
    try:
        tamanho_layout = str(max(
            c["pos_fin"] for c in campos_entrada if c.get("pos_fin")
        ))
    except (ValueError, TypeError):
        tamanho_layout = id_norm.get("tamanholayout", "")
    if tamanho_layout:
        ET.SubElement(root_el, "TamanhoLayout").text = tamanho_layout
    id_evento_val = id_norm.get("identificadorevento", "")
    if id_evento_val:
        ET.SubElement(root_el, "IdentificadorEvento").text = id_evento_val

    # NomeTabela global: lida da aba "Identificação Evento" e aplicada a todos os campos
    nome_tabela_global = id_norm.get("nometabela", "")

    campos_el = ET.SubElement(root_el, "Campos")

    for c in campos_pers:
        raw = c.get("_raw", {})
        rn  = {_norm_aba(k): v for k, v in raw.items()}
        item = ET.SubElement(campos_el, "CampoPersistencia")

        def _add(tag, *keys):
            for k in keys:
                val = rn.get(_norm_aba(k), "")
                if val:
                    ET.SubElement(item, tag).text = str(val)
                    return

        # NomeTabela: valor global da aba Identificação Evento (igual para todos os campos)
        nome_tab = nome_tabela_global or rn.get("nometabela", "")
        if nome_tab:
            ET.SubElement(item, "NomeTabela").text = nome_tab
        _add("NomeColuna",       "NomeColuna")

        # ValorPadrao apenas se não vazio
        vp = rn.get("valorpadrao", "") or c.get("valor_padrao", "")
        if vp:
            ET.SubElement(item, "ValorPadrao").text = vp

        _add("AlinhamentoCampo", "AlinhamentoCampo", "Alinhamento")
        _add("IdentificadorCampo", "IdentificadorCampo")
        _add("NomeCampo",        "NomeCampo")
        _add("DescricaoCampo",   "DescricaoCampo", "Descricao")
        _add("TipoCampo",        "TipoCampo", "Tipo")
        _add("CampoObrigatorio", "CampoObrigatorio", "Obrigatorio")

        # TamanhoCampo apenas se não vazio
        tam = rn.get("tamanhocampo", "") or (str(c["tamanho"]) if c.get("tamanho") else "")
        if tam:
            ET.SubElement(item, "TamanhoCampo").text = tam

    raw_xml = ET.tostring(root_el, encoding="unicode")
    return minidom.parseString(raw_xml).toprettyxml(indent="\t")


def construir_xml_mapa_atributo(dados_por_aba, filepath=None):
    """
    Gera XML attributeMap (namespace ns2) a partir dos campos com MapaAtributo=S.
    defaultValueDefinition lido da aba 'Rule Attribute Valor Padrão'.

    Estrutura:
      <ns2:attributeMap xmlns:ns2="http://rule.saf.cpqd.com.br/">
        <defaultValueDefinition>
          <defaultValueItem dataType="..." pattern="..." value="..."/>
        </defaultValueDefinition>
        <input>
          <origin name="ENRICHMENT">
            <attribute>
              <eventAttribute name="..." type="..."/>
              <ruleAttribute name="..." type="..."/>
              <description>...</description>
              <documentation>...</documentation>
            </attribute>
          </origin>
        </input>
      </ns2:attributeMap>
    """
    NS = "http://rule.saf.cpqd.com.br/"
    # "ns2" é reservado internamente pelo ElementTree (Python 3.12+).
    # Usamos um prefixo seguro e renomeamos no resultado final.
    _NS_PREFIX = "cpqdns"
    ET.register_namespace(_NS_PREFIX, NS)

    campos_entrada = _aba_campos_entrada(dados_por_aba)
    campos_mapa = [
        c for c in campos_entrada
        if _raw_flag(c.get("_raw", {}), "MapaAtributo")
    ]

    default_values = _ler_rule_attribute_valores(filepath) if filepath else []

    root_el = ET.Element(f"{{{NS}}}attributeMap")

    # defaultValueDefinition
    dv_el = ET.SubElement(root_el, "defaultValueDefinition")
    for dv in default_values:
        dv_n = {_norm_aba(k): v for k, v in dv.items()}
        attribs = {}
        for src_key, xml_key in [("datatype", "dataType"), ("pattern", "pattern"), ("value", "value")]:
            val = dv_n.get(src_key, "")
            if val:
                attribs[xml_key] = val
        if attribs:
            ET.SubElement(dv_el, "defaultValueItem", attribs)

    # input → agrupado por Origin
    input_el = ET.SubElement(root_el, "input")
    origins = {}
    for c in campos_mapa:
        raw = c.get("_raw", {})
        rn  = {_norm_aba(k): v for k, v in raw.items()}
        origin = rn.get("origin", "") or rn.get("origem", "") or "UNKNOWN"
        origins.setdefault(origin, []).append(c)

    for origin_name, origin_campos in origins.items():
        origin_el = ET.SubElement(input_el, "origin", {"name": origin_name})
        for c in origin_campos:
            raw = c.get("_raw", {})
            rn  = {_norm_aba(k): v for k, v in raw.items()}
            attr_el = ET.SubElement(origin_el, "attribute")

            event_attr = rn.get("eventattribute", "") or c.get("nome", "")
            rule_attr  = rn.get("ruleattribute",  "") or c.get("nome", "")
            type_val   = rn.get("type", "") or "STRING"
            desc       = (rn.get("description", "")
                          or rn.get("descricaocampo", "")
                          or c.get("descricao", ""))
            doc        = rn.get("documentation", "") or desc

            ea = ET.SubElement(attr_el, "eventAttribute")
            ea.set("name", event_attr)
            ea.set("type", type_val)
            ra = ET.SubElement(attr_el, "ruleAttribute")
            ra.set("name", rule_attr)
            ra.set("type", type_val)
            ET.SubElement(attr_el, "description").text  = desc
            ET.SubElement(attr_el, "documentation").text = doc

    # Usa indentador próprio (minidom falha com prefixos de namespace).
    # Renomeia o prefixo interno → ns2 para corresponder ao gabarito.
    _indent_et(root_el)
    raw_xml = ET.tostring(root_el, encoding="unicode")
    raw_xml = raw_xml.replace(f"{_NS_PREFIX}:", "ns2:")
    raw_xml = raw_xml.replace(f"xmlns:{_NS_PREFIX}=", "xmlns:ns2=")
    return '<?xml version="1.0" ?>\n' + raw_xml


def construir_xml_enriquecimento(dados_por_aba):
    """
    Gera XML DadoExterno (Enriquecimento) a partir das abas:
      'Enriquecimento', 'Enr_ChaveAcesso', 'Enr_CampoRetornado'.
    CDATA é aplicado nos elementos ComandoSQL e SQLChave.

    Estrutura:
      <DadoExterno>
        <Metrica ligado="S" modo="JMX"/>
        <DadoAcesso>
          <ComandoSQL><![CDATA[...]]></ComandoSQL>
          ...
          <GrupoChave><ChaveAcesso>...</ChaveAcesso></GrupoChave>
          ...
          <CampoRetornado>...</CampoRetornado>
        </DadoAcesso>
      </DadoExterno>
    """
    def _find_campos(patts):
        # Exact match first, then endswith — evita "Persistencia_enriquecimento"
        # ser capturado antes de "Enriquecimento"
        for nome, info in dados_por_aba.items():
            n = _norm_aba(nome)
            for p in patts:
                if n == p:
                    return info.get("campos", [])
        for nome, info in dados_por_aba.items():
            n = _norm_aba(nome)
            for p in patts:
                if n.endswith(p):
                    return info.get("campos", [])
        return []

    def _id_enr(rn):
        """Normaliza IdentificadorEnriquecimento para string int (1.0 → '1')."""
        v = rn.get("identificadorenriquecimento", "")
        try:
            return str(int(float(v)))
        except (ValueError, TypeError):
            return str(v).strip()

    enr_campos   = _find_campos(["enriquecimento"])
    chave_campos = _find_campos(["enrchaveacesso", "chaveacesso"])
    camp_campos  = _find_campos(["enrcamporetornado", "camporetornado"])

    # TamanhoTransacao = PosicaoFinal do último campo de Campos Entrada
    campos_entrada_enr = _aba_campos_entrada(dados_por_aba)
    try:
        tamanho_transacao = str(max(
            c["pos_fin"] for c in campos_entrada_enr if c.get("pos_fin")
        ))
    except (ValueError, TypeError):
        tamanho_transacao = ""

    # Indexa por IdentificadorEnriquecimento (chave de ligação entre abas)
    chaves_por_id = {}
    for c in chave_campos:
        rn = {_norm_aba(k): v for k, v in c.get("_raw", {}).items()}
        chaves_por_id.setdefault(_id_enr(rn), []).append(rn)

    retornados_por_id = {}
    for c in camp_campos:
        rn = {_norm_aba(k): v for k, v in c.get("_raw", {}).items()}
        retornados_por_id.setdefault(_id_enr(rn), []).append(rn)

    root_el = ET.Element("DadoExterno")
    ET.SubElement(root_el, "Metrica", {"ligado": "S", "modo": "JMX"})

    def _te(parent, tag, val):
        if val:
            ET.SubElement(parent, tag).text = str(val)

    for c in enr_campos:
        rn   = {_norm_aba(k): v for k, v in c.get("_raw", {}).items()}
        nome = rn.get("nome", "") or c.get("nome", "")
        enr_id = _id_enr(rn)
        da   = ET.SubElement(root_el, "DadoAcesso")

        # ComandoSQL — CDATA aplicado pós-geração
        ET.SubElement(da, "ComandoSQL").text = rn.get("comandosql", "")

        _te(da, "Nome",        nome)
        _te(da, "Descricao",   rn.get("descricao", "") or c.get("descricao", ""))
        _te(da, "TamanhoTransacao", tamanho_transacao or rn.get("tamanhotransacao", ""))
        _te(da, "PersistirEnriquecimento",
            rn.get("persistirenriquecimento", "") or "S")
        _te(da, "PermiteAtualizarSeExistirCache",
            rn.get("permiteatualizarseexistircache", "") or "N")
        _te(da, "OrigemEnriquecimento",
            rn.get("origemenriquecimento", "") or "BD")

        # SQLChave — CDATA aplicado pós-geração
        ET.SubElement(da, "SQLChave").text = rn.get("sqlchave", "")

        # GrupoChave — ligação por IdentificadorEnriquecimento
        grupo = ET.SubElement(da, "GrupoChave")
        for chave_n in chaves_por_id.get(enr_id, []):
            chave_el = ET.SubElement(grupo, "ChaveAcesso")
            _te(chave_el, "Identificador",  chave_n.get("identificador", ""))
            _te(chave_el, "ConversorChave", chave_n.get("conversorchave", ""))
            _te(chave_el, "PosInicial",
                chave_n.get("posinicial", "") or chave_n.get("posicaoinicial", ""))
            _te(chave_el, "PosFinal",
                chave_n.get("posfinal", "") or chave_n.get("posicaofinal", ""))

        _te(da, "DataSource", rn.get("datasource", ""))
        _te(da, "PermiteAtualizarCache",
            rn.get("permiteatualizarcache", "") or "N")

        # CampoRetornado — ligação por IdentificadorEnriquecimento
        for cr_n in retornados_por_id.get(enr_id, []):
            cr = ET.SubElement(da, "CampoRetornado")
            _te(cr, "AliasCampo", cr_n.get("aliascampo", ""))

            # CampoDestino: sempre presente; auto-fechado se vazio
            cd = ET.SubElement(cr, "CampoDestino")
            v = cr_n.get("campodestino", "")
            if v:
                cd.text = v

            _te(cr, "NomeCampo", cr_n.get("nomecampo", ""))
            _te(cr, "TipoCampo", cr_n.get("tipocampo", ""))

            # MascaraCampo: sempre presente; auto-fechado se vazio
            mc = ET.SubElement(cr, "MascaraCampo")
            v = cr_n.get("mascaracampo", "")
            if v:
                mc.text = v

            _te(cr, "PosInicial",
                cr_n.get("posinicial", "") or cr_n.get("posicaoinicial", ""))
            _te(cr, "PosFinal",
                cr_n.get("posfinal", "") or cr_n.get("posicaofinal", ""))
            _te(cr, "MapaDestino", cr_n.get("mapadestino", ""))

        # Campos finais do DadoAcesso — ficam após todos os CampoRetornado
        _te(da, "QuantidadeThreadsInicializacao",
            rn.get("quantidadethreadsinicializacao", ""))
        _te(da, "Prioridade",          rn.get("prioridade", ""))
        _te(da, "PreencherComBrancos", rn.get("preenchercombrancos", ""))

    raw_xml = ET.tostring(root_el, encoding="unicode")
    pretty  = minidom.parseString(raw_xml).toprettyxml(indent="\t")
    # Gabarito exige encoding="UTF-8" na declaração XML
    pretty  = pretty.replace(
        '<?xml version="1.0" ?>',
        '<?xml version="1.0" encoding="UTF-8"?>'
    )

    # Envolve ComandoSQL e SQLChave em CDATA
    def _to_cdata(xml_text, tag):
        def repl(m):
            content = html.unescape(m.group(1))
            return f"<{tag}><![CDATA[{content}]]></{tag}>"
        return re.sub(rf"<{tag}>(.*?)</{tag}>", repl, xml_text, flags=re.DOTALL)

    pretty = _to_cdata(pretty, "ComandoSQL")
    pretty = _to_cdata(pretty, "SQLChave")
    return pretty


# ─────────────────────────────────────────────────────────────────────────────
# Geração de Comandos SQL (ComandosSQL → ComandoSQL.sql)
# ─────────────────────────────────────────────────────────────────────────────

def gerar_comandos_sql(dados_por_aba, filepath=None):
    """
    Gera scripts SQL para os campos com Persistência=S.

    Estrutura do resultado:
      1. Cabeçalhos fixos lidos da aba 'ComandosSQL' do xlsx (linhas onde
         col1 != 'insert na tabela column_configuration')
      2. Um INSERT INTO COLUMN_CONFIGURATION por campo com Persistência=S

    Mapeamento TipoCampo → SQL type:
      TEXTO          → VARCHAR2  (NR_DATA_LENGTH=tamanho, PRECISION=null, SCALE=null)
      DATA / DATA_HORA → DATE    (all null)
      INTEIRO / ID / FK / DECIMAL / NUMERO / NUMBER → NUMBER
                                 (NR_DATA_LENGTH=null, PRECISION=tamanho, SCALE=null)
      outros         → VARCHAR2  (default)
    """
    linhas_sql = []

    # 1. Cabeçalhos fixos do xlsx ─────────────────────────────────────────────
    if filepath:
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = None
            for name in wb.sheetnames:
                if _norm_aba(name) == "comandossql":
                    sheet = wb[name]
                    break
            if sheet is not None:
                for row in sheet.iter_rows(min_row=1, values_only=True):
                    label = str(row[0]).strip() if row[0] is not None else ""
                    sql   = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                    if label.lower() == "insert na tabela column_configuration":
                        continue
                    if sql:
                        linhas_sql.append(sql)
        except Exception:
            pass

    # 2. INSERTs por campo ────────────────────────────────────────────────────
    # NomeTabela via aba "Identificação Evento"
    id_evt = _ler_identificacao_evento(filepath) if filepath else {}
    id_norm = {_norm_aba(k): v for k, v in id_evt.items()}
    nome_tabela = id_norm.get("nometabela", "")

    # Campos com Persistência=S
    campos_entrada = _aba_campos_entrada(dados_por_aba)
    campos_pers = [
        c for c in campos_entrada
        if _raw_flag(c.get("_raw", {}), "Persistência", "Persistencia")
    ]

    _TIPO_NUMBER = {"inteiro", "id", "fk", "decimal", "numero", "number"}
    _TIPO_DATE   = {"data", "data_hora"}

    for c in campos_pers:
        raw  = c.get("_raw", {})
        rn   = {_norm_aba(k): v for k, v in raw.items()}

        nome_coluna  = (rn.get("nomecampo") or c.get("nome") or "").strip()
        descricao    = (rn.get("descricaocampo") or rn.get("descricao") or "").strip()
        # Escapar aspas simples na descrição
        descricao    = descricao.replace("'", "''")

        tipo_campo   = (rn.get("tipocampo") or rn.get("tipo") or "").strip().lower()
        tamanho_raw  = rn.get("tamanho") or c.get("tamanho") or ""
        try:
            tamanho = int(float(str(tamanho_raw)))
        except (ValueError, TypeError):
            tamanho = 0

        nullable_raw = (rn.get("obrigatorio") or rn.get("nullable") or "N").strip().upper()
        # IN_NULLABLE: se obrigatório=S → 0, se não → 1
        in_nullable  = 0 if nullable_raw == "S" else 1

        if tipo_campo in _TIPO_DATE:
            sql_type          = "DATE"
            nr_data_length    = "null"
            nr_data_precision = "null"
            nr_data_scale     = "null"
        elif tipo_campo in _TIPO_NUMBER:
            sql_type          = "NUMBER"
            nr_data_length    = "null"
            nr_data_precision = str(tamanho) if tamanho else "null"
            nr_data_scale     = "null"
        else:
            # VARCHAR2 (TEXTO e demais)
            sql_type          = "VARCHAR2"
            nr_data_length    = str(tamanho) if tamanho else "null"
            nr_data_precision = "null"
            nr_data_scale     = "null"

        insert = (
            f"insert into COLUMN_CONFIGURATION "
            f"(ID_COLUMN_CONFIGURATION,ID_TABLE_CONFIGURATION,ID_DATA_TYPE,"
            f"NM_COLUMN_CONFIGURATION,DS_COLUMN_CONFIGURATION,"
            f"NR_DATA_LENGTH,NR_DATA_PRECISION,NR_DATA_SCALE,IN_NULLABLE,"
            f"IN_PK,IN_FK) values (\n"
            f"  seq_COLUMN_CONFIGURATION.nextval,\n"
            f"  (select ID_TABLE_CONFIGURATION from TABLE_CONFIGURATION "
            f"where NM_TABLE_CONFIGURATION='{nome_tabela}'),\n"
            f"  (select ID_DATA_TYPE from DATA_TYPE "
            f"where NM_DATA_TYPE='{sql_type}'),\n"
            f"  '{nome_coluna}','{descricao}',"
            f"{nr_data_length},{nr_data_precision},{nr_data_scale},"
            f"{in_nullable},0,0);"
        )
        linhas_sql.append(insert)

    return "\n\n".join(linhas_sql)


# ─────────────────────────────────────────────────────────────────────────────
# Validação
# ─────────────────────────────────────────────────────────────────────────────

def validar_campos(campos):
    """Retorna (erros, avisos, infos) com os resultados da validação."""
    erros, avisos, infos = [], [], []

    ativos = [
        c for c in campos
        if (c.get("entrada", "S") or "S").upper() == "S"
        and c.get("pos_ini") and c.get("tamanho")
    ]

    if not ativos:
        return erros, ["Nenhum campo ativo com posição definida."], infos

    ordenados = sorted(ativos, key=lambda c: c["pos_ini"])
    total = sum(c["tamanho"] for c in ordenados)

    # 1. Fórmula PosicaoFinal
    for c in ordenados:
        esperado = c["pos_ini"] + c["tamanho"] - 1
        if c.get("pos_fin") and c["pos_fin"] != esperado:
            erros.append(
                f"Campo '{c['nome']}': PosicaoFinal={c['pos_fin']} "
                f"mas esperado {esperado} (PosIni={c['pos_ini']} + Tam={c['tamanho']} - 1)"
            )

    # 2. Começa em 1
    if ordenados[0]["pos_ini"] != 1:
        avisos.append(
            f"Layout não começa em 1. Primeiro campo '{ordenados[0]['nome']}' "
            f"inicia em {ordenados[0]['pos_ini']}."
        )

    # 3. Continuidade
    for i in range(1, len(ordenados)):
        ant, atu = ordenados[i - 1], ordenados[i]
        prox_esperado = ant["pos_ini"] + ant["tamanho"]
        if atu["pos_ini"] > prox_esperado:
            avisos.append(
                f"GAP entre '{ant['nome']}' (term. {prox_esperado-1}) "
                f"e '{atu['nome']}' (inicia {atu['pos_ini']}) "
                f"— {atu['pos_ini'] - prox_esperado} byte(s)."
            )
        elif atu["pos_ini"] < prox_esperado:
            erros.append(
                f"SOBREPOSIÇÃO: '{ant['nome']}' e '{atu['nome']}' "
                f"se sobrepõem em pos={atu['pos_ini']}."
            )

    # 4. Obrigatórios sem valor
    for c in ordenados:
        if (c.get("obrigatorio") or "").upper() == "S":
            v = (c.get("valor") or c.get("valor_padrao") or "").strip()
            if not v:
                avisos.append(f"Campo obrigatório '{c['nome']}' sem valor preenchido.")

    infos.append(f"Campos de entrada: {len(ordenados)}")
    infos.append(f"Soma dos tamanhos: {total} bytes")
    if ordenados:
        infos.append(f"Posição final do layout: {ordenados[-1]['pos_ini'] + ordenados[-1]['tamanho'] - 1}")

    return erros, avisos, infos


# ─────────────────────────────────────────────────────────────────────────────
# Janela de carregamento (loading)
# ─────────────────────────────────────────────────────────────────────────────

class JanelaCarregando(tk.Toplevel):
    """Diálogo modal com barra de progresso indeterminada exibido durante carregamento."""

    def __init__(self, parent, mensagem="Carregando planilha..."):
        super().__init__(parent)
        self.title("Aguarde")
        self.resizable(False, False)
        self.transient(parent)
        self.protocol("WM_DELETE_WINDOW", lambda: None)  # impede fechar

        self._evento_cancel = threading.Event()

        frm = tk.Frame(self, bg=COR_BG, padx=40, pady=28)
        frm.pack(fill=tk.BOTH, expand=True)

        tk.Label(frm, text="⏳", bg=COR_BG, font=("Segoe UI", 32)).pack()
        self._lbl_msg = tk.Label(
            frm, text=mensagem, bg=COR_BG,
            font=FONT_BOLD, fg="#333333", wraplength=300, justify=tk.CENTER
        )
        self._lbl_msg.pack(pady=(10, 10))

        self._bar = ttk.Progressbar(frm, mode="indeterminate", length=280)
        self._bar.pack()
        self._bar.start(12)

        self._lbl_timer = tk.Label(
            frm, text="00:00", bg=COR_BG,
            font=("Segoe UI", 10), fg="#888888"
        )
        self._lbl_timer.pack(pady=(8, 6))

        tk.Button(
            frm, text="✕  Cancelar",
            command=self._solicitar_cancelamento,
            bg="#c62828", fg="white",
            font=FONT_NORMAL, relief=tk.FLAT,
            padx=14, pady=5, cursor="hand2"
        ).pack()

        # Centraliza sobre o parent
        self.update_idletasks()
        px = parent.winfo_rootx() + parent.winfo_width() // 2 - self.winfo_width() // 2
        py = parent.winfo_rooty() + parent.winfo_height() // 2 - self.winfo_height() // 2
        self.geometry(f"+{px}+{py}")

        self.grab_set()

        self._inicio = time.time()
        self._timer_id = None
        self._tick()

    def _tick(self):
        try:
            elapsed = int(time.time() - self._inicio)
            mins, secs = divmod(elapsed, 60)
            self._lbl_timer.config(text=f"{mins:02d}:{secs:02d}")
            self._timer_id = self.after(1000, self._tick)
        except Exception:
            pass

    def _solicitar_cancelamento(self):
        self._evento_cancel.set()
        try:
            self._lbl_msg.config(text="Cancelando...\nAguarde o passo atual finalizar.")
        except Exception:
            pass

    @property
    def cancelado(self):
        return self._evento_cancel.is_set()

    def atualizar(self, msg):
        """Atualiza a mensagem exibida durante o loading."""
        try:
            self._lbl_msg.config(text=msg)
            self.update_idletasks()
        except Exception:
            pass

    def fechar(self):
        try:
            if self._timer_id:
                self.after_cancel(self._timer_id)
            self._bar.stop()
            self.grab_release()
            self.destroy()
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# Janela de edição de campo (Toplevel)
# ─────────────────────────────────────────────────────────────────────────────

class JanelaEditarCampo(tk.Toplevel):
    TIPOS = ["TEXTO", "INTEIRO", "INTEIRO_LONGO", "DECIMAL", "DATA", "DATA_HORA", "HORA", "BOOLEANO"]
    ALINHAMENTOS = ["", "BRANCO_ESQUERDA", "BRANCO_DIREITA", "ZERO_ESQUERDA", "ZERO_DIREITA"]
    ENTRADAS = ["S", "N"]
    OBRIGATORIOS = ["", "S", "N"]

    def __init__(self, parent, campo=None, on_confirmar=None):
        super().__init__(parent)
        self.title("Editar Campo" if campo else "Novo Campo")
        self.resizable(False, False)
        self.grab_set()

        self.on_confirmar = on_confirmar
        self.resultado = None

        self._vars = {}
        self._build_ui(campo or {})

        self.transient(parent)
        self.wait_window(self)

    def _build_ui(self, campo):
        pad = {"padx": 8, "pady": 3}

        frame = tk.Frame(self, bg=COR_BG, padx=16, pady=12)
        frame.pack(fill=tk.BOTH, expand=True)

        linhas = [
            ("Entrada:",     "entrada",    ttk.Combobox, {"values": self.ENTRADAS, "width": 6}),
            ("Nome:",        "nome",       tk.Entry,     {"width": 32}),
            ("Descrição:",   "descricao",  tk.Entry,     {"width": 32}),
            ("Tipo:",        "tipo",       ttk.Combobox, {"values": self.TIPOS, "width": 18}),
            ("Tamanho:",     "tamanho",    tk.Entry,     {"width": 10}),
            ("Pos. Inicial:","pos_ini",    tk.Entry,     {"width": 10}),
            ("Pos. Final:",  "pos_fin",    tk.Entry,     {"width": 10}),
            ("Alinhamento:", "alinhamento",ttk.Combobox, {"values": self.ALINHAMENTOS, "width": 20}),
            ("Obrigatório:", "obrigatorio",ttk.Combobox, {"values": self.OBRIGATORIOS, "width": 6}),
            ("Valor Padrão:","valor_padrao",tk.Entry,    {"width": 24}),
            ("Valor (XML):", "valor",      tk.Entry,     {"width": 24}),
            ("Coluna DB:",   "coluna_db",  tk.Entry,     {"width": 24}),
        ]

        for i, (label, key, WidgetClass, kw) in enumerate(linhas):
            tk.Label(frame, text=label, bg=COR_BG, font=FONT_NORMAL,
                     anchor=tk.W).grid(row=i, column=0, sticky=tk.W, **pad)

            var = tk.StringVar()
            self._vars[key] = var

            w_kw = dict(kw)
            if WidgetClass == ttk.Combobox:
                w = WidgetClass(frame, textvariable=var, **w_kw)
                w.configure(state="readonly" if key in ("entrada", "tipo", "alinhamento", "obrigatorio") else "normal")
            else:
                w = WidgetClass(frame, textvariable=var, font=FONT_NORMAL, **w_kw)

            # Pos Final read-only (calculado)
            if key == "pos_fin":
                w.configure(state="readonly")

            w.grid(row=i, column=1, sticky=tk.W, **pad)

        frame.columnconfigure(1, weight=1)

        # Auto-calcular Pos Final
        for k in ("tamanho", "pos_ini"):
            self._vars[k].trace_add("write", self._calc_pos_fin)

        # Preenche com valores existentes
        mapa_default = {"entrada": "S", "tipo": "TEXTO"}
        for key, var in self._vars.items():
            val = campo.get(key)
            if val is not None and str(val).strip():
                var.set(str(val))
            elif key in mapa_default:
                var.set(mapa_default[key])

        # Botões
        btn_frame = tk.Frame(self, bg=COR_BG, pady=8)
        btn_frame.pack()

        btn_ok = tk.Button(btn_frame, text="Confirmar", font=FONT_BOLD,
                           bg=COR_BTN_VERDE, fg=COR_BRANCO, relief=tk.FLAT,
                           padx=12, pady=4, command=self._confirmar)
        btn_ok.pack(side=tk.LEFT, padx=6)

        btn_cancel = tk.Button(btn_frame, text="Cancelar", font=FONT_NORMAL,
                               bg=COR_BTN_CINZA, fg=COR_BRANCO, relief=tk.FLAT,
                               padx=12, pady=4, command=self.destroy)
        btn_cancel.pack(side=tk.LEFT, padx=6)

        self.bind("<Return>", lambda e: self._confirmar())
        self.bind("<Escape>", lambda e: self.destroy())

    def _calc_pos_fin(self, *_):
        try:
            pos = int(self._vars["pos_ini"].get())
            tam = int(self._vars["tamanho"].get())
            self._vars["pos_fin"].set(str(pos + tam - 1))
        except ValueError:
            pass

    def _confirmar(self):
        nome = self._vars["nome"].get().strip()
        if not nome:
            messagebox.showwarning("Aviso", "Nome do campo é obrigatório.", parent=self)
            return

        try:
            tamanho = int(self._vars["tamanho"].get()) if self._vars["tamanho"].get() else None
            pos_ini = int(self._vars["pos_ini"].get()) if self._vars["pos_ini"].get() else None
        except ValueError:
            messagebox.showerror("Erro", "Tamanho e Pos. Inicial devem ser números inteiros.", parent=self)
            return

        self.resultado = {
            "entrada":    self._vars["entrada"].get() or "S",
            "nome":       nome,
            "descricao":  self._vars["descricao"].get().strip(),
            "tipo":       self._vars["tipo"].get() or "TEXTO",
            "tamanho":    tamanho,
            "pos_ini":    pos_ini,
            "pos_fin":    (pos_ini + tamanho - 1) if (pos_ini and tamanho) else None,
            "alinhamento":  self._vars["alinhamento"].get(),
            "obrigatorio":  self._vars["obrigatorio"].get(),
            "valor_padrao": self._vars["valor_padrao"].get().strip(),
            "valor":        self._vars["valor"].get(),
            "coluna_db":    self._vars["coluna_db"].get().strip(),
        }

        if self.on_confirmar:
            self.on_confirmar(self.resultado)
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
# Janela de seleção múltipla de campos da origem
# ─────────────────────────────────────────────────────────────────────────────

class JanelaCopiarCampos(tk.Toplevel):
    """
    Diálogo de seleção de campos da planilha origem.
    Exibe uma aba por sheet (estilo Excel) com Listbox de campos em cada uma.
    Aceita dados_por_aba: {nome_aba: {"campos": list, "headers": list}}
    """

    def __init__(self, parent, dados_por_aba, on_confirmar=None):
        super().__init__(parent)
        self.title("Copiar Campos da Origem")
        self.geometry("520x560")
        self.resizable(True, True)
        self.minsize(400, 400)
        self.grab_set()
        self.transient(parent)

        self._dados_por_aba = dados_por_aba
        self._aba_ativa = ""
        self._listboxes: dict = {}          # nome_aba → Listbox
        self._campos_filtrados: dict = {}   # nome_aba → lista filtrada
        self.on_confirmar = on_confirmar

        self._build_ui()
        self.wait_window(self)

    def _build_ui(self):
        frame = tk.Frame(self, bg=COR_BG, padx=10, pady=10)
        frame.pack(fill=tk.BOTH, expand=True)

        # Filtro
        filt_frame = tk.Frame(frame, bg=COR_BG)
        filt_frame.pack(fill=tk.X, pady=(0, 6))
        tk.Label(filt_frame, text="🔍 Filtrar:", bg=COR_BG, font=FONT_NORMAL).pack(side=tk.LEFT)
        self._var_filtro = tk.StringVar()
        self._var_filtro.trace_add("write", lambda *_: self._filtrar())
        tk.Entry(filt_frame, textvariable=self._var_filtro, font=FONT_NORMAL, width=28,
                 relief=tk.FLAT, highlightthickness=1,
                 highlightbackground="#bbb").pack(side=tk.LEFT, padx=4)
        tk.Button(filt_frame, text="✕", command=lambda: self._var_filtro.set(""),
                  bg="#ddd", relief=tk.FLAT, font=FONT_NORMAL).pack(side=tk.LEFT)

        tk.Label(frame, text="Selecione os campos (Ctrl+Click ou Shift+Click para múltiplos):",
                 bg=COR_BG, font=FONT_NORMAL, anchor=tk.W).pack(fill=tk.X, pady=(0, 4))

        # Notebook com uma aba por sheet
        self._nb = ttk.Notebook(frame)
        self._nb.pack(fill=tk.BOTH, expand=True)
        self._nb.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        for nome_aba, aba_info in self._dados_por_aba.items():
            tab_frame = tk.Frame(self._nb, bg=COR_BG, padx=4, pady=4)
            self._nb.add(tab_frame, text=f"  {nome_aba}  ")
            lb = self._criar_listbox(tab_frame)
            self._listboxes[nome_aba] = lb
            self._campos_filtrados[nome_aba] = list(aba_info.get("campos", []))

        # Seleciona a primeira aba
        nomes = list(self._dados_por_aba.keys())
        if nomes:
            self._aba_ativa = nomes[0]
            self._nb.select(0)

        # Botões de seleção rápida + contador
        sel_frame = tk.Frame(frame, bg=COR_BG)
        sel_frame.pack(fill=tk.X, pady=(6, 0))
        tk.Button(sel_frame, text="Selecionar Todos", command=self._sel_todos,
                  bg="#eceff1", font=FONT_NORMAL, relief=tk.FLAT).pack(side=tk.LEFT, padx=2)
        tk.Button(sel_frame, text="Limpar Seleção", command=self._sel_nenhum,
                  bg="#eceff1", font=FONT_NORMAL, relief=tk.FLAT).pack(side=tk.LEFT, padx=2)
        self._lbl_sel = tk.Label(sel_frame, text="0 selecionado(s)",
                                  bg=COR_BG, fg="#555", font=FONT_NORMAL)
        self._lbl_sel.pack(side=tk.RIGHT)

        # Botões de ação
        btn_frame = tk.Frame(self, bg=COR_BG, pady=8)
        btn_frame.pack()
        self._btn_copiar = tk.Button(btn_frame, text="⬇ Copiar 0 campos", font=FONT_BOLD,
                                      bg=COR_BTN_LARANJA, fg=COR_BRANCO, relief=tk.FLAT,
                                      padx=12, pady=4, command=self._confirmar,
                                      state=tk.DISABLED)
        self._btn_copiar.pack(side=tk.LEFT, padx=6)
        tk.Button(btn_frame, text="Cancelar", font=FONT_NORMAL,
                  bg=COR_BTN_CINZA, fg=COR_BRANCO, relief=tk.FLAT,
                  padx=12, pady=4, command=self.destroy).pack(side=tk.LEFT, padx=6)

        self.bind("<Escape>", lambda e: self.destroy())
        self._filtrar()

    def _criar_listbox(self, parent):
        """Cria e retorna um Listbox com scrollbar dentro do parent."""
        list_frame = tk.Frame(parent, bg=COR_BG)
        list_frame.pack(fill=tk.BOTH, expand=True)
        lb = tk.Listbox(list_frame, selectmode=tk.EXTENDED, font=FONT_NORMAL,
                        activestyle="dotbox", exportselection=False,
                        relief=tk.FLAT, highlightthickness=1, highlightbackground="#bbb")
        vsb = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=lb.yview)
        lb.configure(yscrollcommand=vsb.set)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        lb.bind("<<ListboxSelect>>", self._on_select)
        return lb

    def _on_tab_changed(self, _evt=None):
        """Atualiza aba ativa e reseta o contador de seleção."""
        nomes = list(self._dados_por_aba.keys())
        tabs = self._nb.tabs()
        if not tabs:
            return
        idx = self._nb.index("current")
        if idx < len(nomes):
            self._aba_ativa = nomes[idx]
        self._on_select()

    def _get_lb(self):
        """Retorna o Listbox da aba atualmente selecionada."""
        return self._listboxes.get(self._aba_ativa)

    def _filtrar(self):
        """Aplica filtro de texto em todas as abas."""
        filtro = self._var_filtro.get().lower()
        for nome_aba, aba_info in self._dados_por_aba.items():
            campos = aba_info.get("campos", [])
            filtrados = [
                c for c in campos
                if not filtro
                   or filtro in (c.get("nome") or "").lower()
                   or filtro in (c.get("descricao") or "").lower()
            ]
            self._campos_filtrados[nome_aba] = filtrados
            lb = self._listboxes.get(nome_aba)
            if not lb:
                continue
            lb.delete(0, tk.END)
            for c in filtrados:
                nome   = c.get("nome", "")
                extras = []
                if c.get("tipo"):    extras.append(c["tipo"])
                if c.get("tamanho"): extras.append(f"{c['tamanho']}b")
                label = f"{nome}  [{', '.join(extras)}]" if extras else nome
                lb.insert(tk.END, label)
        self._on_select()

    def _sel_todos(self):
        lb = self._get_lb()
        if lb:
            lb.select_set(0, tk.END)
        self._on_select()

    def _sel_nenhum(self):
        lb = self._get_lb()
        if lb:
            lb.selection_clear(0, tk.END)
        self._on_select()

    def _on_select(self, _evt=None):
        lb = self._get_lb()
        n = len(lb.curselection()) if lb else 0
        self._lbl_sel.config(text=f"{n} selecionado(s)")
        plural = "s" if n != 1 else ""
        self._btn_copiar.config(
            text=f"⬇ Copiar {n} campo{plural}",
            state=tk.NORMAL if n > 0 else tk.DISABLED
        )

    def _confirmar(self):
        lb = self._get_lb()
        if not lb:
            self.destroy()
            return
        indices = lb.curselection()
        filtrados = self._campos_filtrados.get(self._aba_ativa, [])
        resultado = [filtrados[i] for i in indices]
        if self.on_confirmar:
            self.on_confirmar(resultado)
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
# Aplicação principal
# ─────────────────────────────────────────────────────────────────────────────

class GeradorXMLApp:

    # ── Inicialização ────────────────────────────────────────────────────────

    def __init__(self, root):
        self.root = root
        self.root.title("Gerador de XML — Planilhas de Eventos")
        self.root.geometry("1280x820")
        self.root.configure(bg=COR_BG)
        self.root.minsize(900, 600)

        self._campos: list = []             # campos da aba ativa
        self._dados_por_aba: dict = {}      # nome_aba → {"campos": list, "headers": list, "sections": dict}
        self._aba_ativa: str = ""           # aba atualmente exibida
        self._headers_ativos: list = []     # headers da aba ativa (ordem original da planilha)
        self._sections_ativos: dict = {}    # seções de metadados da aba ativa (para construir_xml)
        self._dados_por_aba_origem: dict = {}  # nome_aba → {"campos": list, "headers": list, "sections": dict}
        self._arquivo_principal = None
        self._arquivo_origem = None
        self._path_principal_pendente = None   # selecionado mas ainda não carregado
        self._path_origem_pendente    = None
        self._idx_editando = -1             # índice do campo sendo editado

        # Widgets do notebook de abas (criados em _build_tabela)
        self._nb_abas = None                # ttk.Notebook da planilha principal
        self._trees_abas: dict = {}         # nome_aba → Treeview
        self._tree = None                   # Treeview da aba atualmente ativa
        self._ignorar_tab_change = False    # evita recursão ao selecionar aba

        # Widgets de preview XML (criados em _build_painel_direito)
        self._txt_xmls: dict = {}           # key → tk.Text  (LayoutEntrada, LayoutPersistencia, mapaAtributo, DadoExterno)
        self._txt_xml  = None               # alias para _txt_xmls["LayoutEntrada"] (compatibilidade)

        self._setup_estilos()
        self._build_ui()
        self._bind_atalhos()

    def _setup_estilos(self):
        s = ttk.Style()
        s.theme_use("clam")
        s.configure("Treeview", rowheight=24, font=FONT_NORMAL, background=COR_BRANCO)
        s.configure("Treeview.Heading", font=FONT_BOLD)
        s.map("Treeview", background=[("selected", "#bbdefb")])

    # ── Construção da UI ─────────────────────────────────────────────────────

    def _build_ui(self):
        self._build_menubar()
        self._build_toolbar()
        self._build_conteudo()
        self._build_statusbar()

    def _build_menubar(self):
        mb = tk.Menu(self.root)
        self.root.config(menu=mb)

        m_arq = tk.Menu(mb, tearoff=0)
        m_arq.add_command(label="Selecionar Principal...  Ctrl+O", command=self.selecionar_principal)
        m_arq.add_command(label="Selecionar Origem...",            command=self.selecionar_origem)
        m_arq.add_command(label="Carregar Planilhas",              command=self.carregar_planilhas)
        m_arq.add_separator()
        m_arq.add_command(label="Salvar Planilha  Ctrl+S",        command=self.salvar_planilha)
        m_arq.add_separator()
        m_arq.add_command(label="Sair",                           command=self.root.quit)
        mb.add_cascade(label="Arquivo", menu=m_arq)

        m_fer = tk.Menu(mb, tearoff=0)
        m_fer.add_command(label="Validar Soma  F5",               command=self.validar)
        m_fer.add_command(label="Recalcular Posições",            command=self.recalcular_posicoes)
        m_fer.add_separator()
        m_fer.add_command(label="Gerar XMLs (todos)  F6",         command=self.gerar_xml)
        m_fer.add_command(label="Pré-visualizar XML  F7",         command=self.preview_xml)
        mb.add_cascade(label="Ferramentas", menu=m_fer)

    def _btn(self, parent, text, command, bg, **pack_kw):
        b = tk.Button(parent, text=text, command=command, bg=bg,
                      fg=COR_BRANCO, relief=tk.FLAT, font=FONT_NORMAL,
                      padx=9, pady=4, cursor="hand2",
                      activebackground=bg, activeforeground=COR_BRANCO)
        b.pack(**pack_kw)
        return b

    def _build_toolbar(self):
        bar = tk.Frame(self.root, bg=COR_TOOLBAR, pady=5)
        bar.pack(fill=tk.X)

        # ── Planilha principal ───────────────────────────────────────────────
        frp = tk.LabelFrame(bar, text="Planilha Principal", bg=COR_TOOLBAR,
                            font=FONT_NORMAL, padx=4, pady=2)
        frp.pack(side=tk.LEFT, padx=8)

        self._btn(frp, "📂 Selecionar Principal", self.selecionar_principal,
                  COR_BTN_VERDE, side=tk.LEFT, padx=2)
        self._lbl_principal = tk.Label(frp, text="—", bg=COR_TOOLBAR,
                                       fg="#555", font=("Segoe UI", 8))
        self._lbl_principal.pack(side=tk.LEFT, padx=6)

        # ── Planilha origem ──────────────────────────────────────────────────
        fro = tk.LabelFrame(bar, text="Planilha Origem", bg=COR_TOOLBAR,
                            font=FONT_NORMAL, padx=4, pady=2)
        fro.pack(side=tk.LEFT, padx=4)

        self._btn(fro, "📂 Selecionar Origem", self.selecionar_origem,
                  COR_BTN_AZUL, side=tk.LEFT, padx=2)
        self._lbl_origem = tk.Label(fro, text="—", bg=COR_TOOLBAR,
                                    fg="#555", font=("Segoe UI", 8))
        self._lbl_origem.pack(side=tk.LEFT, padx=6)

        # ── Carregar (dispara o loading de todas as planilhas selecionadas) ──
        frl = tk.Frame(bar, bg=COR_TOOLBAR)
        frl.pack(side=tk.LEFT, padx=6)
        self._btn_carregar = self._btn(
            frl, "⬇ Carregar Planilhas", self.carregar_planilhas,
            COR_BTN_VERDE, side=tk.LEFT, padx=2
        )
        self._btn_carregar.configure(state=tk.DISABLED)

        # ── Copiar campos ─────────────────────────────────────────────────────
        frc = tk.LabelFrame(bar, text="Copiar Campos da Origem", bg=COR_TOOLBAR,
                            font=FONT_NORMAL, padx=4, pady=2)
        frc.pack(side=tk.LEFT, padx=4)

        self._btn_copiar_origem = self._btn(frc, "⬇ Copiar Campos...", self.copiar_campo,
                  COR_BTN_LARANJA, side=tk.LEFT, padx=2)
        self._btn_copiar_origem.configure(state=tk.DISABLED)

        # ── Ações ────────────────────────────────────────────────────────────
        fra = tk.Frame(bar, bg=COR_TOOLBAR)
        fra.pack(side=tk.RIGHT, padx=8)

        self._btn(fra, "📄 Gerar XMLs [F6]",       self.gerar_xml,         COR_BTN_ROXO,  side=tk.RIGHT, padx=2)
        self._btn(fra, "👁 Preview XML [F7]",      self.preview_xml,       COR_BTN_CINZA, side=tk.RIGHT, padx=2)
        self._btn(fra, "✔ Validar [F5]",           self.validar,           COR_BTN_CINZA, side=tk.RIGHT, padx=2)
        self._btn(fra, "💾 Salvar Planilha",        self.salvar_planilha,   COR_BTN_TEAL,  side=tk.RIGHT, padx=2)

    def _build_conteudo(self):
        paned = tk.PanedWindow(self.root, orient=tk.HORIZONTAL,
                               sashwidth=6, bg="#cfd8dc")
        paned.pack(fill=tk.BOTH, expand=True, padx=6, pady=4)

        # ── Esquerda: tabela ─────────────────────────────────────────────────
        frame_esq = tk.Frame(paned, bg=COR_BG)
        paned.add(frame_esq, width=810)
        self._build_tabela(frame_esq)

        # ── Direita: detalhes / validação / XML ──────────────────────────────
        frame_dir = tk.Frame(paned, bg=COR_BG)
        paned.add(frame_dir, width=400)
        self._build_painel_direito(frame_dir)

    def _build_tabela(self, parent):
        # Barra de filtro
        filt = tk.Frame(parent, bg=COR_BG)
        filt.pack(fill=tk.X, pady=(2, 4))

        tk.Label(filt, text="🔍 Filtrar:", bg=COR_BG, font=FONT_NORMAL).pack(side=tk.LEFT)
        self._var_filtro = tk.StringVar()
        self._var_filtro.trace_add("write", lambda *_: self._atualizar_tabela())
        tk.Entry(filt, textvariable=self._var_filtro, font=FONT_NORMAL, width=28,
                 relief=tk.FLAT, highlightthickness=1,
                 highlightbackground="#bbb").pack(side=tk.LEFT, padx=4)
        tk.Button(filt, text="✕", command=lambda: self._var_filtro.set(""),
                  bg="#ddd", relief=tk.FLAT, font=FONT_NORMAL).pack(side=tk.LEFT)

        # Notebook de abas (uma por planilha)
        self._nb_abas = ttk.Notebook(parent)
        self._nb_abas.pack(fill=tk.BOTH, expand=True)
        self._nb_abas.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        # Barra de ações da tabela
        bbar = tk.Frame(parent, bg=COR_BG, pady=4)
        bbar.pack(fill=tk.X)

        self._btn(bbar, "+ Novo",       self.novo_campo,       COR_BTN_VERDE,  side=tk.LEFT, padx=2)
        self._btn(bbar, "✎ Editar",    self.editar_campo,     COR_BTN_AZUL,   side=tk.LEFT, padx=2)
        self._btn(bbar, "🗑 Remover",   self.remover_campo,    COR_BTN_VERM,   side=tk.LEFT, padx=2)
        self._btn(bbar, "⟳ Recalcular",self.recalcular_posicoes, COR_BTN_LARANJA, side=tk.LEFT, padx=2)

        self._lbl_count = tk.Label(bbar, text="0 campos", bg=COR_BG,
                                   fg="#555", font=FONT_NORMAL)
        self._lbl_count.pack(side=tk.RIGHT, padx=8)

    def _build_painel_direito(self, parent):
        nb = ttk.Notebook(parent)
        nb.pack(fill=tk.BOTH, expand=True)

        # Aba Validação
        frv = tk.Frame(nb, bg=COR_BG, padx=4, pady=4)
        nb.add(frv, text="  Validação  ")
        self._build_aba_validacao(frv)

        # Abas de XML Preview — uma por tipo de XML gerado
        _XML_ABAS = [
            ("LayoutEntrada",      "  LayoutEntrada  "),
            ("LayoutPersistencia", "  LayoutPersistencia  "),
            ("mapaAtributo",       "  mapaAtributo  "),
            ("DadoExterno",        "  DadoExterno  "),
            ("ComandoSQL",         "  ComandoSQL  "),
        ]
        for key, label in _XML_ABAS:
            frx = tk.Frame(nb, bg=COR_BG, padx=4, pady=4)
            nb.add(frx, text=label)
            self._txt_xmls[key] = self._build_xml_tab(frx, key)

        # Alias de compatibilidade
        self._txt_xml = self._txt_xmls["LayoutEntrada"]
        self._notebook = nb

    def _build_aba_validacao(self, parent):
        self._btn(parent, "▶ Validar Agora [F5]", self.validar,
                  COR_BTN_CINZA, anchor=tk.NW, pady=(0, 6))

        self._txt_val = tk.Text(parent, font=FONT_MONO, wrap=tk.WORD,
                                state=tk.DISABLED, bg="#fafafa",
                                relief=tk.FLAT, highlightthickness=1,
                                highlightbackground="#ccc")
        vsb = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=self._txt_val.yview)
        self._txt_val.configure(yscrollcommand=vsb.set)

        self._txt_val.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # Tags de cor para o texto
        self._txt_val.tag_configure("ok",     foreground="#2e7d32", font=FONT_BOLD)
        self._txt_val.tag_configure("erro",   foreground="#c62828", font=FONT_BOLD)
        self._txt_val.tag_configure("aviso",  foreground="#e65100")
        self._txt_val.tag_configure("info",   foreground="#1565c0")
        self._txt_val.tag_configure("titulo", font=FONT_BOLD)

    def _build_xml_tab(self, parent, key):
        """Cria aba de preview XML com botão de atualização. Retorna o widget Text."""
        self._btn(parent, "🔄 Atualizar Preview", self._preview_todas_abas,
                  COR_BTN_ROXO, anchor=tk.NW, pady=(0, 6))

        frm = tk.Frame(parent)
        frm.pack(fill=tk.BOTH, expand=True)

        txt = tk.Text(frm, font=FONT_MONO, wrap=tk.NONE,
                      bg="#1e1e1e", fg="#d4d4d4",
                      insertbackground="white",
                      relief=tk.FLAT)
        vsb = ttk.Scrollbar(frm, orient=tk.VERTICAL,   command=txt.yview)
        hsb = ttk.Scrollbar(frm, orient=tk.HORIZONTAL, command=txt.xview)
        txt.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        txt.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frm.rowconfigure(0, weight=1)
        frm.columnconfigure(0, weight=1)

        txt.tag_configure("tag",   foreground="#569cd6")
        txt.tag_configure("attr",  foreground="#9cdcfe")
        txt.tag_configure("value", foreground="#ce9178")
        return txt

    def _build_statusbar(self):
        bar = tk.Frame(self.root, bg="#37474f", height=26)
        bar.pack(fill=tk.X, side=tk.BOTTOM)
        bar.pack_propagate(False)

        self._var_status = tk.StringVar(value="Pronto. Carregue uma planilha para começar.")
        tk.Label(bar, textvariable=self._var_status, bg="#37474f", fg=COR_BRANCO,
                 font=FONT_NORMAL, anchor=tk.W, padx=10).pack(side=tk.LEFT, fill=tk.X, expand=True)

        self._var_total = tk.StringVar()
        tk.Label(bar, textvariable=self._var_total, bg="#37474f", fg="#90a4ae",
                 font=FONT_NORMAL, padx=10).pack(side=tk.RIGHT)

    # ── Atalhos ──────────────────────────────────────────────────────────────

    def _bind_atalhos(self):
        self.root.bind("<Control-o>", lambda _: self.selecionar_principal())
        self.root.bind("<Control-s>", lambda _: self.salvar_planilha())
        self.root.bind("<F5>",        lambda _: self.validar())
        self.root.bind("<F6>",        lambda _: self.gerar_xml())
        self.root.bind("<F7>",        lambda _: self.preview_xml())
        self.root.bind("<Delete>",    lambda _: self.remover_campo())

    # ── Carregar planilhas ────────────────────────────────────────────────────

    def _executar_em_thread(self, tarefa, on_sucesso, on_erro, mensagem="Carregando..."):
        """
        Executa `tarefa()` em thread separada enquanto exibe JanelaCarregando.
        Chama on_sucesso(resultado) ou on_erro(excecao) na thread principal via after().
        """
        janela = JanelaCarregando(self.root, mensagem)

        def _runner():
            try:
                resultado = tarefa()
                self.root.after(0, lambda: _finalizar(resultado, None))
            except Exception as exc:
                self.root.after(0, lambda e=exc: _finalizar(None, e))

        def _finalizar(resultado, erro):
            janela.fechar()
            if erro:
                on_erro(erro)
            else:
                on_sucesso(resultado)

        threading.Thread(target=_runner, daemon=True).start()

    def selecionar_principal(self):
        """Abre diálogo de seleção — apenas armazena o caminho, não carrega ainda."""
        path = filedialog.askopenfilename(
            title="Selecionar Planilha Principal",
            filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("Todos", "*.*")]
        )
        if not path:
            return
        self._path_principal_pendente = path
        nome = os.path.basename(path)
        self._lbl_principal.config(text=f"● {nome}", fg="#e65100")
        self._btn_carregar.configure(state=tk.NORMAL)
        self._set_status(f"Principal selecionada: {nome}  —  clique em 'Carregar Planilhas' para carregar.")

    def selecionar_origem(self):
        """Abre diálogo de seleção — apenas armazena o caminho, não carrega ainda."""
        path = filedialog.askopenfilename(
            title="Selecionar Planilha Origem",
            filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("Todos", "*.*")]
        )
        if not path:
            return
        self._path_origem_pendente = path
        nome = os.path.basename(path)
        self._lbl_origem.config(text=f"● {nome}", fg="#e65100")
        self._btn_carregar.configure(state=tk.NORMAL)
        self._set_status(f"Origem selecionada: {nome}  —  clique em 'Carregar Planilhas' para carregar.")

    def _aplicar_principal(self, path, dados):
        """Aplica os dados da planilha principal já carregados na UI."""
        self._dados_por_aba = dados
        self._arquivo_principal = path
        nomes = list(dados.keys())
        aba_padrao = next(
            (n for n in nomes if _normalizar_chave(n) in ("camposentrada",)),
            nomes[0]
        )
        self._reconstruir_abas_principal(dados, aba_padrao)
        nome = os.path.basename(path)
        self._lbl_principal.config(text=nome, fg="#1565c0")
        total = sum(len(v["campos"]) for v in dados.values())
        self._set_status(f"Principal carregada: {nome}  —  {len(dados)} aba(s), {total} campos")

    def _aplicar_origem(self, path, dados):
        """Aplica os dados da planilha origem já carregados na UI."""
        self._dados_por_aba_origem = dados
        self._arquivo_origem = path
        self._btn_copiar_origem.configure(state=tk.NORMAL)
        nome = os.path.basename(path)
        self._lbl_origem.config(text=nome, fg="#2e7d32")
        total = sum(len(v["campos"]) for v in dados.values())
        self._set_status(
            f"Origem carregada: {nome}  —  {len(dados)} aba(s), {total} campos disponíveis"
        )

    def carregar_planilhas(self):
        """Carrega em sequência todas as planilhas com seleção pendente."""
        arquivos = []
        if self._path_principal_pendente:
            arquivos.append(("principal", self._path_principal_pendente))
        if self._path_origem_pendente:
            arquivos.append(("origem", self._path_origem_pendente))

        if not arquivos:
            messagebox.showwarning("Aviso", "Selecione pelo menos uma planilha antes de carregar.")
            return

        total = len(arquivos)
        janela = JanelaCarregando(self.root, "Preparando...")

        def _runner():
            resultados = {}
            erro_info  = None
            for i, (tipo, path) in enumerate(arquivos):
                if janela.cancelado:
                    break
                nome = os.path.basename(path)
                msg  = f"Carregando:\n{nome}\n\narquivo {i + 1} de {total}"
                self.root.after(0, lambda m=msg: janela.atualizar(m))
                try:
                    dados = ler_todas_abas(path)
                    if tipo == "origem" and not dados:
                        campos = self._ler_xlsx_generico(path)
                        nome_fb = os.path.splitext(os.path.basename(path))[0]
                        dados = {nome_fb: {"campos": campos, "headers": []}}
                    resultados[tipo] = (path, dados)
                except Exception as e:
                    erro_info = (tipo, path, e)
                    break
            cancelado = janela.cancelado
            self.root.after(0, lambda: _finalizar(resultados, erro_info, cancelado))

        def _finalizar(resultados, erro, cancelado):
            janela.fechar()

            if cancelado:
                # Rollback: nenhum dado é aplicado — estado anterior preservado
                self._set_status("Carregamento cancelado — nenhuma alteração aplicada.")
                return

            if erro:
                tipo, path, exc = erro
                messagebox.showerror(
                    "Erro ao carregar",
                    f"{os.path.basename(path)}:\n{exc}"
                )

            if "principal" in resultados:
                path, dados = resultados["principal"]
                if not dados:
                    messagebox.showwarning(
                        "Aviso",
                        f"{os.path.basename(path)}: nenhuma aba com campos detectada."
                    )
                else:
                    self._aplicar_principal(path, dados)
                    self._path_principal_pendente = None

            if "origem" in resultados:
                path, dados = resultados["origem"]
                self._aplicar_origem(path, dados)
                self._path_origem_pendente = None

            # Desabilita "Carregar" se não há mais nada pendente
            if not self._path_principal_pendente and not self._path_origem_pendente:
                self._btn_carregar.configure(state=tk.DISABLED)

            # Mensagem final consolidada na status bar
            partes = []
            if "principal" in resultados and resultados["principal"][1]:
                partes.append(f"Principal: {os.path.basename(resultados['principal'][0])}")
            if "origem" in resultados:
                partes.append(f"Origem: {os.path.basename(resultados['origem'][0])}")
            if partes:
                self._set_status("Carregadas — " + "  |  ".join(partes))

        threading.Thread(target=_runner, daemon=True).start()

    # ── Abas estilo Excel (notebook) ──────────────────────────────────────────

    def _criar_tree_aba(self, parent, headers):
        """Cria um Treeview com as colunas originais da aba e retorna o widget."""
        cols = headers if headers else ["NomeCampo"]

        wrap = tk.Frame(parent, bg=COR_BG)
        wrap.pack(fill=tk.BOTH, expand=True)

        tree = ttk.Treeview(wrap, columns=cols, show="headings", selectmode="browse")
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=120, anchor=tk.W, minwidth=50)

        vsb = ttk.Scrollbar(wrap, orient=tk.VERTICAL,   command=tree.yview)
        hsb = ttk.Scrollbar(wrap, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)

        tree.tag_configure("par",   background="#f5f9ff")
        tree.tag_configure("impar", background=COR_BRANCO)
        tree.tag_configure("erro",  background="#ffebee")
        tree.tag_configure("aviso", background="#fff8e1")

        tree.bind("<<TreeviewSelect>>", self._on_selecionar)
        tree.bind("<Double-1>",         self._on_duplo_clique)

        return tree

    def _reconstruir_abas_principal(self, dados, aba_padrao):
        """Recria todas as abas do Notebook com os dados carregados."""
        self._ignorar_tab_change = True

        # Limpa abas e trees anteriores
        for tab in self._nb_abas.tabs():
            self._nb_abas.forget(tab)
        self._trees_abas.clear()

        nomes = list(dados.keys())
        for nome_aba in nomes:
            aba_info = dados[nome_aba]
            headers = aba_info.get("headers", [])
            campos  = aba_info.get("campos", [])

            frame = tk.Frame(self._nb_abas, bg=COR_BG)
            self._nb_abas.add(frame, text=f"  {nome_aba}  ")

            tree = self._criar_tree_aba(frame, headers)
            self._trees_abas[nome_aba] = tree

            # Popula com os dados brutos
            for i, c in enumerate(campos):
                raw    = c.get("_raw", {})
                vals   = tuple(raw.get(h, "") for h in headers) if headers else (c.get("nome", ""),)
                pos_ini = c.get("pos_ini")
                tam     = c.get("tamanho")
                pos_fin = c.get("pos_fin")
                if pos_ini and tam and pos_fin and pos_ini + tam - 1 != pos_fin:
                    tag = "erro"
                elif not pos_ini or not tam:
                    tag = "aviso"
                else:
                    tag = "par" if i % 2 == 0 else "impar"
                tree.insert("", tk.END, iid=str(i), tags=(tag,), values=vals)

        # Seleciona a aba padrão
        idx_padrao = nomes.index(aba_padrao) if aba_padrao in nomes else 0
        self._ignorar_tab_change = False
        self._nb_abas.select(idx_padrao)
        # Dispara manualmente (select pode não disparar evento se já estava naquele tab)
        self._on_tab_changed()

    def _on_tab_changed(self, _evt=None):
        """Chamado quando o usuário clica em outra aba do notebook."""
        if self._ignorar_tab_change or not self._nb_abas.tabs():
            return
        idx = self._nb_abas.index("current")
        nomes = list(self._dados_por_aba.keys())
        if idx >= len(nomes):
            return
        nome_aba = nomes[idx]
        self._aba_ativa = nome_aba
        aba = self._dados_por_aba.get(nome_aba, {})
        self._campos = aba.get("campos", [])
        self._headers_ativos = aba.get("headers", [])
        self._sections_ativos = aba.get("sections", {})
        self._tree = self._trees_abas.get(nome_aba)
        self._lbl_count.config(text=f"{len(self._campos)} campos")
        self._atualizar_total()
        nome_xml = _nome_xml_para_aba(nome_aba)
        self._set_status(f"Aba: {nome_aba}  ({len(self._campos)} campos)  →  {nome_xml}")

    def _mudar_aba(self, nome_aba):
        """Seleciona a aba pelo nome (usado internamente)."""
        nomes = list(self._dados_por_aba.keys())
        if nome_aba in nomes:
            self._nb_abas.select(nomes.index(nome_aba))
            self._on_tab_changed()

    def _ler_xlsx_generico(self, path):
        """Lê a primeira aba como lista de campos, usando a 1ª linha como cabeçalho."""
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        headers = [_cell_str(c.value, f"Col{i}") for i, c in enumerate(ws[1])]
        campos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            nome = _cell_str(row[0]) if row else ""
            if not nome:
                continue
            c = {"nome": nome, "entrada": "S", "tipo": "TEXTO"}
            for h, v in zip(headers[1:], row[1:]):
                c[_normalizar_chave(h)] = _cell_str(v)
            campos.append(c)
        return campos

    # ── Copiar campo da origem ────────────────────────────────────────────────

    def copiar_campo(self):
        if not self._dados_por_aba_origem:
            messagebox.showwarning("Aviso", "Carregue a planilha origem primeiro.")
            return

        def _processar(campos_selecionados):
            if not campos_selecionados:
                return

            novos, duplicatas = [], []
            for orig in campos_selecionados:
                nome = orig.get("nome", "")
                existente = next((c for c in self._campos if c.get("nome") == nome), None)
                if existente:
                    duplicatas.append((existente, orig))
                else:
                    novos.append(orig)

            # ── Índices de dados específicos da origem (por NomeCampo normalizado) ─
            pers_idx = {}   # abas "Persistenc*"
            mapa_idx = {}   # abas "RuleAttribute*" ou "MapaAtributo*"
            for nome_aba, info in self._dados_por_aba_origem.items():
                n_aba = _norm_aba(nome_aba)
                for idx, patts in [(pers_idx, ("persistenc",)),
                                   (mapa_idx, ("ruleattribute", "mapaatributo", "attributemap"))]:
                    if any(n_aba.startswith(p) for p in patts):
                        for c in info.get("campos", []):
                            raw_p = c.get("_raw", {})
                            rn_p  = {_norm_aba(k): v for k, v in raw_p.items()}
                            nc = rn_p.get("nomecampo", "") or c.get("nome", "")
                            if nc:
                                idx.setdefault(_norm_aba(nc), raw_p)
                        break

            # ── NomeTabela do arquivo principal ───────────────────────────────────
            id_evt_principal = (
                _ler_identificacao_evento(self._arquivo_principal)
                if self._arquivo_principal else {}
            )
            id_norm_principal = {_norm_aba(k): v for k, v in id_evt_principal.items()}
            nome_tabela_principal = id_norm_principal.get("nometabela", "")

            # ── Helpers ──────────────────────────────────────────────────────────
            def _id_reservado(v):
                return (1000 <= v < 2000) or (20000 <= v <= 21000)

            def _set_raw(raw, val, std_key, *norm_alts):
                for k in list(raw.keys()):
                    if _norm_aba(k) in norm_alts:
                        raw[k] = str(val)
                        return
                raw[std_key] = str(val)

            _SKIP_MERGE = {
                "persistencia", "persistência",
                "mapaatributo",
                "posicaoinicial", "posinicial",
                "posicaofinal", "posfinal",
                "identificadorcampo",
            }

            def _mesclar_persistencia(raw, nome_campo):
                """Mescla dados de persistência da origem e força NomeTabela principal."""
                pers_raw = pers_idx.get(_norm_aba(nome_campo), {})
                for k, v in pers_raw.items():
                    if _norm_aba(k) not in _SKIP_MERGE:
                        raw.setdefault(k, v)
                # NomeTabela sempre vem do arquivo principal
                if nome_tabela_principal:
                    _set_raw(raw, nome_tabela_principal, "NomeTabela", "nometabela")

            def _mesclar_mapa(raw, nome_campo):
                """Mescla dados de mapa de atributo da origem."""
                mapa_raw = mapa_idx.get(_norm_aba(nome_campo), {})
                for k, v in mapa_raw.items():
                    if _norm_aba(k) not in _SKIP_MERGE:
                        raw.setdefault(k, v)

            # ── Próximo IdentificadorCampo sequencial ─────────────────────────────
            used_ids = []
            for c in self._campos:
                try:
                    v = int(float(c.get("id", "")))
                    if not _id_reservado(v):
                        used_ids.append(v)
                except (ValueError, TypeError):
                    pass
            next_id = (max(used_ids) + 1) if used_ids else 1

            # ── Adiciona campos novos ─────────────────────────────────────────────
            for orig in novos:
                novo      = dict(orig)
                novo["_raw"] = dict(orig.get("_raw", {}))
                raw       = novo["_raw"]

                # Se Persistência=S, mescla dados do tab de persistência da origem
                if _raw_flag(raw, "Persistência", "Persistencia"):
                    _mesclar_persistencia(raw, novo.get("nome", ""))
                # Se MapaAtributo=S, mescla dados de mapa de atributo da origem
                if _raw_flag(raw, "MapaAtributo"):
                    _mesclar_mapa(raw, novo.get("nome", ""))

                # Posição sequencial da planilha principal
                ativos = [c for c in self._campos if c.get("pos_ini") and c.get("tamanho")]
                prox   = (max(c["pos_ini"] + c["tamanho"] for c in ativos) if ativos else 1)
                novo["pos_ini"] = prox
                if novo.get("tamanho"):
                    novo["pos_fin"] = prox + novo["tamanho"] - 1
                _set_raw(raw, prox, "PosicaoInicial", "posicaoinicial", "posinicial")
                if novo.get("pos_fin"):
                    _set_raw(raw, novo["pos_fin"], "PosicaoFinal", "posicaofinal", "posfinal")

                # IdentificadorCampo sequencial da planilha principal
                novo["id"] = str(next_id)
                _set_raw(raw, next_id, "IdentificadorCampo", "identificadorcampo")

                self._campos.append(novo)
                next_id += 1

            # ── Duplicatas ────────────────────────────────────────────────────────
            atualizados = 0
            if duplicatas:
                nomes_dup = "\n".join(f"  • {e.get('nome')}" for e, _ in duplicatas)
                if messagebox.askyesno("Campos já existem",
                        f"Os seguintes campos já existem na planilha principal:\n{nomes_dup}\n\n"
                        "Deseja atualizar os atributos deles (tipo, tamanho, alinhamento etc.)?"):
                    for existente, orig in duplicatas:
                        for key in ("tipo", "tamanho", "alinhamento", "descricao",
                                    "obrigatorio", "coluna_db", "valor_padrao"):
                            if orig.get(key) is not None:
                                existente[key] = orig[key]
                        if not existente.get("valor"):
                            existente["valor"] = orig.get("valor_padrao", "")
                        # Atualiza também dados de persistência, mapa de atributo e NomeTabela
                        raw_ex = existente.setdefault("_raw", {})
                        if _raw_flag(raw_ex, "Persistência", "Persistencia"):
                            _mesclar_persistencia(raw_ex, existente.get("nome", ""))
                        if _raw_flag(raw_ex, "MapaAtributo"):
                            _mesclar_mapa(raw_ex, existente.get("nome", ""))
                        atualizados += 1

            self._atualizar_tabela()
            partes = []
            if novos:
                partes.append(f"{len(novos)} copiado(s)")
            if atualizados:
                partes.append(f"{atualizados} atualizado(s)")
            if partes:
                self._set_status(f"Campos da origem: {', '.join(partes)}.")

        JanelaCopiarCampos(self.root, self._dados_por_aba_origem, on_confirmar=_processar)

    # ── Tabela ────────────────────────────────────────────────────────────────

    def _atualizar_tabela(self):
        """Repopula a Treeview da aba ativa com os dados atuais de self._campos."""
        if not self._tree:
            return

        tree    = self._tree
        headers = self._headers_ativos
        filtro  = self._var_filtro.get().lower()
        exib    = 0

        for item in tree.get_children():
            tree.delete(item)

        for i, c in enumerate(self._campos):
            nome = (c.get("nome") or "").lower()
            desc = (c.get("descricao") or "").lower()
            if filtro and filtro not in nome and filtro not in desc:
                continue

            raw  = c.get("_raw", {})
            vals = tuple(raw.get(h, "") for h in headers) if headers else (c.get("nome", ""),)

            pos_ini = c.get("pos_ini")
            tam     = c.get("tamanho")
            pos_fin = c.get("pos_fin")
            if pos_ini and tam and pos_fin and pos_ini + tam - 1 != pos_fin:
                tag = "erro"
            elif not pos_ini or not tam:
                tag = "aviso"
            else:
                tag = "par" if i % 2 == 0 else "impar"

            tree.insert("", tk.END, iid=str(i), tags=(tag,), values=vals)
            exib += 1

        self._lbl_count.config(text=f"{exib}/{len(self._campos)} campos")
        self._atualizar_total()

    def _atualizar_total(self):
        ativos = [c for c in self._campos if c.get("pos_ini") and c.get("tamanho")]
        if ativos:
            total = sum(c["tamanho"] for c in ativos)
            maxi  = max(c["pos_ini"] + c["tamanho"] - 1 for c in ativos)
            self._var_total.set(f"Tamanho total: {total} bytes | Pos. máx: {maxi}")
        else:
            self._var_total.set("")

    def _on_selecionar(self, _evt):
        pass  # pode expandir para pré-preencher detalhe

    def _on_duplo_clique(self, _evt):
        self.editar_campo()

    # ── Ações CRUD ────────────────────────────────────────────────────────────

    def novo_campo(self):
        # Sugere próxima posição
        ativos = [c for c in self._campos if c.get("pos_ini") and c.get("tamanho")]
        prox = (max(c["pos_ini"] + c["tamanho"] for c in ativos) if ativos else 1)
        JanelaEditarCampo(
            self.root,
            campo={"pos_ini": prox, "entrada": "S"},
            on_confirmar=self._adicionar_campo
        )

    def _adicionar_campo(self, campo):
        max_id = max(
            (int(c["id"]) for c in self._campos if str(c.get("id", "")).isdigit()),
            default=0
        )
        campo["id"] = str(max_id + 1)
        self._campos.append(campo)
        self._atualizar_tabela()
        self._set_status(f"Campo '{campo['nome']}' adicionado.")

    def editar_campo(self):
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo("Aviso", "Selecione um campo na tabela.")
            return
        idx = int(sel[0])
        if idx >= len(self._campos):
            return

        def _aplicar(novo):
            novo["id"]    = self._campos[idx].get("id", "")
            novo["linha"] = self._campos[idx].get("linha", "")
            self._campos[idx] = novo
            self._atualizar_tabela()
            self._set_status(f"Campo '{novo['nome']}' atualizado.")

        JanelaEditarCampo(self.root, campo=self._campos[idx], on_confirmar=_aplicar)

    def remover_campo(self):
        sel = self._tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        if idx >= len(self._campos):
            return
        nome = self._campos[idx].get("nome", "?")
        if messagebox.askyesno("Confirmar", f"Remover o campo '{nome}'?"):
            self._campos.pop(idx)
            self._atualizar_tabela()
            self._set_status(f"Campo '{nome}' removido.")

    def recalcular_posicoes(self):
        if not messagebox.askyesno("Recalcular Posições",
                "Recalcula posições de TODOS os campos de entrada em sequência a partir de 1.\n"
                "Continuar?"):
            return
        ativos = sorted(
            [c for c in self._campos if (c.get("entrada","S") or "S").upper()=="S" and c.get("tamanho")],
            key=lambda c: c.get("pos_ini", 99999)
        )
        pos = 1
        for c in ativos:
            c["pos_ini"] = pos
            c["pos_fin"] = pos + c["tamanho"] - 1
            pos += c["tamanho"]
        self._atualizar_tabela()
        self._set_status(f"Posições recalculadas. Total: {pos-1} bytes.")

    # ── Salvar planilha ───────────────────────────────────────────────────────

    def salvar_planilha(self):
        if not self._campos:
            messagebox.showwarning("Aviso", "Nenhum campo para salvar.")
            return

        if self._arquivo_principal:
            # Sempre salva em uma cópia _Novo, preservando o original
            base, ext = os.path.splitext(self._arquivo_principal)
            # Remove sufixo _Novo anterior para não empilhar (ex: Arq_Novo_Novo)
            if base.endswith("_Novo"):
                base = base[:-5]
            path = base + "_Novo" + ext
        else:
            path = filedialog.asksaveasfilename(
                title="Salvar Planilha",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")]
            )
            if not path:
                return

        try:
            ext = os.path.splitext(path)[1].lower()
            if ext in (".xlsx", ".xls"):
                if self._arquivo_principal and self._dados_por_aba:
                    # Salva preservando toda a estrutura original (todos os tabs, formatação)
                    salvar_xlsx_estruturado(
                        self._arquivo_principal, path, self._dados_por_aba
                    )
                else:
                    # Fallback: sem arquivo original, cria planilha simples
                    salvar_xlsx(path, self._campos, self._aba_ativa or "Campos Entrada")
            else:
                salvar_csv(path, self._campos)
            self._set_status(f"Planilha salva: {os.path.basename(path)}")
            messagebox.showinfo("Sucesso", f"Planilha salva:\n{path}")
        except Exception as e:
            messagebox.showerror("Erro ao salvar", str(e))

    # ── Validação ─────────────────────────────────────────────────────────────

    def validar(self):
        if not self._campos:
            messagebox.showwarning("Aviso", "Carregue uma planilha primeiro.")
            return

        erros, avisos, infos = validar_campos(self._campos)

        self._txt_val.configure(state=tk.NORMAL)
        self._txt_val.delete(1.0, tk.END)

        self._txt_val.insert(tk.END, "═" * 42 + "\n", "titulo")
        self._txt_val.insert(tk.END, "   RESULTADO DA VALIDAÇÃO\n", "titulo")
        self._txt_val.insert(tk.END, "═" * 42 + "\n\n", "titulo")

        if infos:
            self._txt_val.insert(tk.END, "INFORMAÇÕES:\n", "info")
            for i in infos:
                self._txt_val.insert(tk.END, f"  • {i}\n", "info")
            self._txt_val.insert(tk.END, "\n")

        if avisos:
            self._txt_val.insert(tk.END, "AVISOS:\n", "aviso")
            for a in avisos:
                self._txt_val.insert(tk.END, f"  ⚠  {a}\n", "aviso")
            self._txt_val.insert(tk.END, "\n")

        if erros:
            self._txt_val.insert(tk.END, "ERROS:\n", "erro")
            for e in erros:
                self._txt_val.insert(tk.END, f"  ✗  {e}\n", "erro")
            self._txt_val.insert(tk.END, "\n")
        else:
            self._txt_val.insert(tk.END, "✔  Sem erros de posicionamento!\n", "ok")

        self._txt_val.configure(state=tk.DISABLED)

        # Navega para aba de validação
        self._notebook.select(0)

        status = f"Validação: {'OK' if not erros else f'{len(erros)} erro(s)'}"
        if avisos:
            status += f" | {len(avisos)} aviso(s)"
        self._set_status(status)

    # ── XML ───────────────────────────────────────────────────────────────────

    # Índice de cada aba XML dentro do notebook (tab 0 = Validação)
    _XML_TAB_KEYS = ["LayoutEntrada", "LayoutPersistencia", "mapaAtributo", "DadoExterno", "ComandoSQL"]

    def _atualizar_tab_xml(self, key, xml_str):
        """Popula a aba de XML preview correspondente sem trocar de aba."""
        txt = self._txt_xmls.get(key)
        if not txt:
            return
        txt.configure(state=tk.NORMAL)
        txt.delete(1.0, tk.END)
        txt.insert(tk.END, xml_str)
        txt.configure(state=tk.DISABLED)

    def _gerar_xml_str(self, key):
        """Gera e retorna o conteúdo (XML ou SQL) para a aba indicada."""
        if key == "LayoutEntrada":
            aba_e = None
            for nome, info in self._dados_por_aba.items():
                if _norm_aba(nome) == "camposentrada":
                    aba_e = info
                    break
            if aba_e is None and self._dados_por_aba:
                aba_e = next(iter(self._dados_por_aba.values()))
            return construir_xml(
                aba_e.get("campos", [])   if aba_e else self._campos,
                aba_e.get("headers", [])  if aba_e else self._headers_ativos,
                "Campos Entrada",
                aba_e.get("sections", {}) if aba_e else self._sections_ativos,
            )
        elif key == "LayoutPersistencia":
            return construir_xml_persistencia(
                self._dados_por_aba, self._arquivo_principal
            )
        elif key == "mapaAtributo":
            return construir_xml_mapa_atributo(
                self._dados_por_aba, self._arquivo_principal
            )
        elif key == "DadoExterno":
            return construir_xml_enriquecimento(self._dados_por_aba)
        elif key == "ComandoSQL":
            return gerar_comandos_sql(self._dados_por_aba, self._arquivo_principal)
        return None

    def _preview_xml_tab(self, key):
        """Gera e exibe o conteúdo de uma aba, selecionando-a no notebook."""
        if not self._dados_por_aba:
            messagebox.showwarning("Aviso", "Carregue uma planilha primeiro.")
            return
        try:
            xml_str = self._gerar_xml_str(key)
            if xml_str is None:
                return
            self._atualizar_tab_xml(key, xml_str)
            self._notebook.select(1 + self._XML_TAB_KEYS.index(key))
        except Exception as e:
            messagebox.showerror(f"Erro ao gerar {key}", str(e))

    def _preview_todas_abas(self):
        """Atualiza o conteúdo de todas as abas de preview em thread separada."""
        if not self._dados_por_aba:
            messagebox.showwarning("Aviso", "Carregue uma planilha primeiro.")
            return

        total = len(self._XML_TAB_KEYS)
        janela = JanelaCarregando(self.root, "Gerando previews...")

        def _runner():
            resultados = {}
            erros = []
            for i, key in enumerate(self._XML_TAB_KEYS):
                if janela.cancelado:
                    break
                msg = f"Gerando preview:\n{key}\n\n{i + 1} de {total}"
                self.root.after(0, lambda m=msg: janela.atualizar(m))
                try:
                    resultados[key] = self._gerar_xml_str(key)
                except Exception as e:
                    erros.append(f"{key}: {e}")
            cancelado = janela.cancelado
            self.root.after(0, lambda: _finalizar(resultados, erros, cancelado))

        def _finalizar(resultados, erros, cancelado):
            janela.fechar()
            if cancelado:
                # Rollback: nenhuma aba é atualizada — conteúdo anterior preservado
                self._set_status("Preview cancelado — conteúdo anterior preservado.")
                return
            for key, xml_str in resultados.items():
                if xml_str is not None:
                    self._atualizar_tab_xml(key, xml_str)
            if erros:
                messagebox.showerror("Erro ao atualizar preview", "\n".join(erros))

        threading.Thread(target=_runner, daemon=True).start()

    def preview_xml(self):
        """Atalho F7: atualiza todas as abas de preview."""
        self._preview_todas_abas()

    def gerar_xml(self):
        if not self._dados_por_aba:
            messagebox.showwarning("Aviso", "Carregue uma planilha primeiro.")
            return

        # Usa "Campos Entrada" como referência para validação
        aba_entrada = None
        for nome, info in self._dados_por_aba.items():
            if _norm_aba(nome) == "camposentrada":
                aba_entrada = info
                break
        if aba_entrada is None and self._dados_por_aba:
            aba_entrada = next(iter(self._dados_por_aba.values()))

        campos_validar = aba_entrada.get("campos", []) if aba_entrada else self._campos
        erros, _, _ = validar_campos(campos_validar)
        if erros:
            if not messagebox.askyesno("Validação com erros",
                    f"Existem {len(erros)} erro(s) de validação.\n"
                    "Deseja gerar os XMLs mesmo assim?"):
                return

        # Pede diretório de saída
        dir_inicial = os.path.dirname(self._arquivo_principal) if self._arquivo_principal else ""
        dir_saida = filedialog.askdirectory(
            title="Selecione o diretório para salvar os XMLs",
            initialdir=dir_inicial
        )
        if not dir_saida:
            return

        gerados      = []
        erros_geracao = []

        # 1. LayoutEntrada.xml
        try:
            xml_str = construir_xml(
                campos_validar,
                aba_entrada.get("headers", [])  if aba_entrada else self._headers_ativos,
                "Campos Entrada",
                aba_entrada.get("sections", {}) if aba_entrada else self._sections_ativos,
            )
            path = os.path.join(dir_saida, "LayoutEntrada.xml")
            with open(path, "w", encoding="utf-8") as f:
                f.write(xml_str)
            gerados.append("LayoutEntrada.xml")
            self._atualizar_tab_xml("LayoutEntrada", xml_str)
        except Exception as e:
            erros_geracao.append(f"LayoutEntrada.xml: {e}")

        # 2. LayoutPersistencia.xml
        try:
            xml_str = construir_xml_persistencia(
                self._dados_por_aba, self._arquivo_principal
            )
            path = os.path.join(dir_saida, "LayoutPersistencia.xml")
            with open(path, "w", encoding="utf-8") as f:
                f.write(xml_str)
            gerados.append("LayoutPersistencia.xml")
            self._atualizar_tab_xml("LayoutPersistencia", xml_str)
        except Exception as e:
            erros_geracao.append(f"LayoutPersistencia.xml: {e}")

        # 3. mapaAtributo.xml
        try:
            xml_str = construir_xml_mapa_atributo(
                self._dados_por_aba, self._arquivo_principal
            )
            path = os.path.join(dir_saida, "mapaAtributo.xml")
            with open(path, "w", encoding="utf-8") as f:
                f.write(xml_str)
            gerados.append("mapaAtributo.xml")
            self._atualizar_tab_xml("mapaAtributo", xml_str)
        except Exception as e:
            erros_geracao.append(f"mapaAtributo.xml: {e}")

        # 4. DadoExterno.xml (Enriquecimento)
        try:
            xml_str = construir_xml_enriquecimento(self._dados_por_aba)
            path = os.path.join(dir_saida, "DadoExterno.xml")
            with open(path, "w", encoding="utf-8") as f:
                f.write(xml_str)
            gerados.append("DadoExterno.xml")
            self._atualizar_tab_xml("DadoExterno", xml_str)
        except Exception as e:
            erros_geracao.append(f"DadoExterno.xml: {e}")

        # 5. ComandoSQL.sql
        try:
            sql_str = gerar_comandos_sql(self._dados_por_aba, self._arquivo_principal)
            path = os.path.join(dir_saida, "ComandoSQL.sql")
            with open(path, "w", encoding="utf-8") as f:
                f.write(sql_str)
            gerados.append("ComandoSQL.sql")
            self._atualizar_tab_xml("ComandoSQL", sql_str)
        except Exception as e:
            erros_geracao.append(f"ComandoSQL.sql: {e}")

        # Resultado
        linhas_ok  = "\n".join(f"  ✔ {n}" for n in gerados)
        linhas_err = "\n".join(f"  ✗ {e}" for e in erros_geracao)
        msg = f"XMLs gerados em:\n{dir_saida}\n\n{linhas_ok}"
        if erros_geracao:
            msg += f"\n\nErros:\n{linhas_err}"
            messagebox.showwarning("XMLs gerados com erros", msg)
        else:
            messagebox.showinfo("Sucesso", msg)

        # Seleciona a aba LayoutEntrada para o usuário ver o resultado
        if gerados:
            self._notebook.select(1)

        self._set_status(
            f"{len(gerados)} XML(s) gerado(s) em: {os.path.basename(dir_saida)}"
            + (f"  |  {len(erros_geracao)} erro(s)" if erros_geracao else "")
        )

    # ── Utilidades ────────────────────────────────────────────────────────────

    def _set_status(self, msg):
        self._var_status.set(msg)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    root = tk.Tk()
    try:
        root.tk.call("tk", "scaling", 1.25)
    except Exception:
        pass
    GeradorXMLApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
